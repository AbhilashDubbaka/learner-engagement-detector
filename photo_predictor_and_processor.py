import numpy as np
import cv2
import dlib
from imutils import face_utils
from head_pose_estimator import Headpose_Estimator
from scipy.spatial import distance as dist
from keras.models import load_model
import openpyxl
from tqdm import tqdm
import os
import glob
import pandas as pd
import csv

from parameters import DATASET, VIDEO_PREDICTOR, OTHER

np.random.seed(OTHER.random_state)

class Photo_Predictor:

    def __init__(self):

        #Load all models
        self.net = cv2.dnn.readNetFromCaffe(DATASET.config_file, DATASET.model_file)
        self.face_landmarks_predictor = dlib.shape_predictor(VIDEO_PREDICTOR.face_landmarks_file)
        self.secondary_face_detector = dlib.get_frontal_face_detector()
        self.model_AU1 = load_model('models/AU1/AU1.h5')
        self.model_AU2a = load_model('models/AU2/AU2a.h5')
        self.model_AU2b = load_model('models/AU2/AU2b.h5')
        self.model_AU4 = load_model('models/AU4/AU4.h5')
        self.model_AU5 = load_model('models/AU5/AU5.h5')
        self.model_AU6 = load_model('models/AU6/AU6.h5')
        self.model_AU9 = load_model('models/AU9/AU9.h5')
        self.model_AU12 = load_model('models/AU12/AU12.h5')
        self.model_AU15 = load_model('models/AU15/AU15.h5')
        self.model_AU25 = load_model('models/AU25/AU25.h5')
        self.model_AU26 = load_model('models/AU26/AU26.h5') 
        (self.lStart, self.lEnd) = face_utils.FACIAL_LANDMARKS_IDXS["left_eye"]
        (self.rStart, self.rEnd) = face_utils.FACIAL_LANDMARKS_IDXS["right_eye"]
        self.hpe = Headpose_Estimator()

    def aggregate_AU_predictions(self, only_face_image):
        AU_predictions_per_frame = []

        image_grayscale = cv2.cvtColor(only_face_image, cv2.COLOR_RGB2GRAY)

        # resize to normalize data size
        whole_face = cv2.resize(image_grayscale, (DATASET.face_horizontal_size, DATASET.face_vertical_size))

        # Obtain specific faces including flipped images
        upper_face = whole_face[0:DATASET.upper_face_vertical_size, ].copy()
        upper_face_flip = cv2.flip(upper_face, 1)
        lower_face = whole_face[115:(115+DATASET.lower_face_vertical_size), ].copy()
        lower_face = cv2.resize(lower_face, (DATASET.face_horizontal_size, DATASET.lower_face_vertical_size))
        lower_face_flip = cv2.flip(lower_face, 1)
        whole_face = cv2.resize(whole_face, (DATASET.face_horizontal_size, DATASET.face_vertical_size))
        whole_face_flip = cv2.flip(whole_face, 1)

        #Reshape all faces for model input
        upper_face = np.array(upper_face).reshape(-1, DATASET.upper_face_vertical_size, DATASET.face_horizontal_size, 1)
        upper_face_flip = np.array(upper_face_flip).reshape(-1, DATASET.upper_face_vertical_size, DATASET.face_horizontal_size, 1)
        whole_face = np.array(whole_face).reshape(-1, DATASET.face_vertical_size, DATASET.face_horizontal_size, 1)
        whole_face_flip = np.array(whole_face_flip).reshape(-1, DATASET.face_vertical_size, DATASET.face_horizontal_size, 1)
        lower_face = np.array(lower_face).reshape(-1, DATASET.lower_face_vertical_size, DATASET.face_horizontal_size, 1)
        lower_face_flip = np.array(lower_face_flip).reshape(-1, DATASET.lower_face_vertical_size, DATASET.face_horizontal_size, 1)
        
        # Scale all images
        upper_face = upper_face/255
        upper_face_flip = upper_face_flip/255
        whole_face = whole_face/255
        whole_face_flip = whole_face_flip/255
        lower_face = lower_face/255
        lower_face_flip = lower_face_flip/255
        
        #Predict for all faces
        AU1a = self.model_AU1.predict(upper_face)
        AU1b = self.model_AU1.predict(upper_face_flip)
        AU1 = [[AU1a[0][0]*0.5 + AU1b[0][0]*0.5, AU1a[0][1]*0.5 + AU1b[0][1]*0.5]]
        AU2a_1 = self.model_AU2a.predict(upper_face)
        AU2a_2 = self.model_AU2a.predict(upper_face_flip)
        AU2a = [[AU2a_1[0][0]*0.5 + AU2a_2[0][0]*0.5, AU2a_1[0][1]*0.5 + AU2a_2[0][1]*0.5]]
        AU2b_1 = self.model_AU2b.predict(upper_face)
        AU2b_2 = self.model_AU2b.predict(upper_face_flip)
        AU2b = [[AU2b_1[0][0]*0.5 + AU2b_2[0][0]*0.5, AU2b_1[0][1]*0.5 + AU2b_2[0][1]*0.5]]
        AU4a = self.model_AU4.predict(upper_face)
        AU4b = self.model_AU4.predict(upper_face_flip)
        AU4 = [[AU4a[0][0]*0.5 + AU4b[0][0]*0.5, AU4a[0][1]*0.5 + AU4b[0][1]*0.5]]
        AU5a = self.model_AU5.predict(upper_face)
        AU5b = self.model_AU5.predict(upper_face_flip)
        AU5 = [[AU5a[0][0]*0.5 + AU5b[0][0]*0.5, AU5a[0][1]*0.5 + AU5b[0][1]*0.5]]
        AU6a = self.model_AU6.predict(whole_face)
        AU6b = self.model_AU6.predict(whole_face_flip)
        AU6 = [[AU6a[0][0]*0.5 + AU6b[0][0]*0.5, AU6a[0][1]*0.5 + AU6b[0][1]*0.5]]
        AU9a = self.model_AU9.predict(whole_face)
        AU9b = self.model_AU9.predict(whole_face_flip)
        AU9 = [[AU9a[0][0]*0.5 + AU9b[0][0]*0.5, AU9a[0][1]*0.5 + AU9b[0][1]*0.5]]
        AU12a = self.model_AU12.predict(lower_face)
        AU12b = self.model_AU12.predict(lower_face_flip)
        AU12 = [[AU12a[0][0]*0.5 + AU12b[0][0]*0.5, AU12a[0][1]*0.5 + AU12b[0][1]*0.5]]
        AU15a = self.model_AU15.predict(lower_face)
        AU15b = self.model_AU15.predict(lower_face_flip)
        AU15 = [[AU15a[0][0]*0.5 + AU15b[0][0]*0.5, AU15a[0][1]*0.5 + AU15b[0][1]*0.5]]
        AU25a = self.model_AU25.predict(lower_face)
        AU25b = self.model_AU25.predict(lower_face_flip)
        AU25 = [[AU25a[0][0]*0.5 + AU25b[0][0]*0.5, AU25a[0][1]*0.5 + AU25b[0][1]*0.5]]
        AU26a = self.model_AU26.predict(lower_face)
        AU26b = self.model_AU26.predict(lower_face_flip)
        AU26 = [[AU26a[0][0]*0.5 + AU26b[0][0]*0.5, AU26a[0][1]*0.5 + AU26b[0][1]*0.5]]
         
        #Append all predictions into list of predictions for the frame
        AU_predictions_per_frame.append(AU1[0][1])
        AU_predictions_per_frame.append(AU2a[0][1])
        AU_predictions_per_frame.append(AU2b[0][1])
        AU_predictions_per_frame.append(AU4[0][1])
        AU_predictions_per_frame.append(AU5[0][1])
        AU_predictions_per_frame.append(AU6[0][1])
        AU_predictions_per_frame.append(AU9[0][1])
        AU_predictions_per_frame.append(AU12[0][1])
        AU_predictions_per_frame.append(AU15[0][1])
        AU_predictions_per_frame.append(AU25[0][1])
        AU_predictions_per_frame.append(AU26[0][1])
        
        return AU_predictions_per_frame

    def rotate(self, image, angle, center=None, scale=1.0):
        # grab the dimensions of the image
        (h, w) = image.shape[:2]
    
        # if the center is None, initialize it as the center of
        # the image
        if center is None:
            center = (w // 2, h // 2)
    
        # perform the rotation
        M = cv2.getRotationMatrix2D(center, angle, scale)
        rotated = cv2.warpAffine(image, M, (w, h))
    
        # return the rotated image
        return rotated
    
    def process_image(self, image_path):
        image_path = image_path.replace('\\', '/')
        image = cv2.imread(image_path)
        # detect faces
        (h, w) = image.shape[:2]
        #Rotate the face so it is aligned horizontally i.e. Roll is at 0 degrees
        angle, shape = self.get_rotation_angle(image)
        non_rotated_image = image.copy()
        image = self.rotate(image, angle)
        blob = cv2.dnn.blobFromImage(cv2.resize(image, (300,300)), 1.0, (300, 300), (104.0, 177.0, 123.0))
        self.net.setInput(blob)
        detections = self.net.forward()
        AU_predictions_per_photo = []
        landmarks_per_photo = []
        
        # loop over the detections and pick highest confidence picture
        detection_index = 0
        max_confidence = 0
        for i in range(0, detections.shape[2]):
            confidence = detections[0, 0, i, 2]
            if max_confidence < confidence:
                max_confidence = confidence
                detection_index = i

        i = detection_index
		# extract the confidence (i.e., probability) associated with the
		# prediction
        confidence = detections[0, 0, i, 2]
		# filter out weak detections by ensuring the `confidence` is
		# greater than the minimum confidence
        if ((confidence <= DATASET.min_confidence) or (detections.shape[2] == 0) or (shape is None)):
            return AU_predictions_per_photo, landmarks_per_photo
		
        # compute the (x, y)-coordinates of the bounding box for the
		# object
        box = detections[0, 0, i, 3:7] * np.array([w, h, w, h])
        (startX, startY, endX, endY) = box.astype("int")
        only_face_image = image[startY-2:endY+2, startX-2:endX+2].copy()
		
        #Send aligned face for landmark detection
        landmarks_per_photo = self.landmark_detection(non_rotated_image, shape)
        
        #Aligned face image to be detected again for face and sent for AU detection
        AU_predictions_per_photo = self.aggregate_AU_predictions(only_face_image)
            
        return AU_predictions_per_photo, landmarks_per_photo

    def landmark_detection(self, image, shape):
        landmarks_per_frame = []
        #Head pose estimations
        angles = self.hpe.process_image(image, shape)
        pitch_min = angles[0]
        pitch_avg = angles[0]
        yaw_min = angles[1]
        yaw_avg = angles[1]
        roll_min = angles[2]
        roll_avg = angles[2]
        landmarks_per_frame.append(pitch_min)
        landmarks_per_frame.append(pitch_avg)
        landmarks_per_frame.append(yaw_min)
        landmarks_per_frame.append(yaw_avg)
        landmarks_per_frame.append(roll_min)
        landmarks_per_frame.append(roll_avg)
        
        #Facial landmarks
        shape = face_utils.shape_to_np(shape)
        landmarks_per_frame.append(shape)
        #Detect drowsiness
        drowsiness_per_frame = self.drowsiness_detection(shape)
        landmarks_per_frame.append(drowsiness_per_frame)
        
        return landmarks_per_frame
        
    def drowsiness_detection(self, shape):
        # extract the left and right eye coordinates, then use the
        # coordinates to compute the eye aspect ratio for both eyes
        leftEye = shape[self.lStart:self.lEnd]
        rightEye = shape[self.rStart:self.rEnd]
        leftEAR = self.eye_aspect_ratio(leftEye)
        rightEAR = self.eye_aspect_ratio(rightEye)

        # average the eye aspect ratio together for both eyes
        ear = (leftEAR + rightEAR) / 2.0

        return ear

    def eye_aspect_ratio(self, eye):
    	# compute the euclidean distances between the two sets of
    	# vertical eye landmarks (x, y)-coordinates
    	A = dist.euclidean(eye[1], eye[5])
    	B = dist.euclidean(eye[2], eye[4])

    	# compute the euclidean distance between the horizontal
    	# eye landmark (x, y)-coordinates
    	C = dist.euclidean(eye[0], eye[3])

    	# compute the eye aspect ratio
    	ear = (A + B) / (2.0 * C)

    	# return the eye aspect ratio
    	return ear

    def get_rotation_angle(self, image):
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        rects = self.secondary_face_detector(gray, 1)
        angle = 0
        landmarks = None
        if len(rects) > 0:      
            # convert the landmark (x, y)-coordinates to a NumPy array
            landmarks = self.face_landmarks_predictor(gray, rects[0])
            shape = face_utils.shape_to_np(landmarks)
    		
            if (len(shape)==68):
    			# extract the left and right eye (x, y)-coordinates
                (lStart, lEnd) = face_utils.FACIAL_LANDMARKS_68_IDXS["left_eye"]
                (rStart, rEnd) = face_utils.FACIAL_LANDMARKS_68_IDXS["right_eye"]
            else:
                (lStart, lEnd) = face_utils.FACIAL_LANDMARKS_5_IDXS["left_eye"]
                (rStart, rEnd) = face_utils.FACIAL_LANDMARKS_5_IDXS["right_eye"]
    			
            leftEyePts = shape[lStart:lEnd]
            rightEyePts = shape[rStart:rEnd]
    
    		# compute the center of mass for each eye
            leftEyeCenter = leftEyePts.mean(axis=0).astype("int")
            rightEyeCenter = rightEyePts.mean(axis=0).astype("int")
    
    		# compute the angle between the eye centroids
            dY = rightEyeCenter[1] - leftEyeCenter[1]
            dX = rightEyeCenter[0] - leftEyeCenter[0]
            angle = np.degrees(np.arctan2(dY, dX)) - 180
        return angle, landmarks

def annotate_CK_with_landmarks(predictor):
    filename = "CK+_landmarks.xlsx"
    tab_name = "CK+_landmarks"
    book = openpyxl.load_workbook(filename)
    sheet = book[tab_name]
    for rownum in tqdm(range(1, len(sheet['A']))):        
        _, landmarks = predictor.process_image(sheet.cell(row=rownum+1, column=1).value)
        if not landmarks:
            continue
        else:
            landmarks_flatten = landmarks[6].tolist()
            for col in range(1, 69):
                k = col * 2
                sheet.cell(row = rownum+1, column = k).value = landmarks_flatten[col-1][0]
                sheet.cell(row = rownum+1, column = k+1).value = landmarks_flatten[col-1][1]
            sheet.cell(row = rownum+1, column = 138).value = landmarks[7]
    book.save(filename)
    
    book = openpyxl.load_workbook(filename)
    #Calculate geometric features
    print("Calculating geometric features...")
    if "geometric_calcs" in book.sheetnames:
            sheet = book["geometric_calcs"]
            book.remove(sheet)
    sheet = book.create_sheet(title="geometric_calcs")
    sheet.cell(row=1, column=1).value = "Image path"
    sheet.cell(row=1, column=2).value = "AU1_l"
    sheet.cell(row=1, column=3).value = "AU1_r"
    sheet.cell(row=1, column=4).value = "AU2_l"
    sheet.cell(row=1, column=5).value = "AU2_r"
    sheet.cell(row=1, column=6).value = "AU4_t"
    sheet.cell(row=1, column=7).value = "AU9_l"
    sheet.cell(row=1, column=8).value = "AU9_r"
    sheet.cell(row=1, column=9).value = "AU12_t"
    sheet.cell(row=1, column=10).value = "AU12_l"
    sheet.cell(row=1, column=11).value = "AU12_r"
    sheet.cell(row=1, column=12).value = "AU15_l"
    sheet.cell(row=1, column=13).value = "AU15_r"
    sheet.cell(row=1, column=14).value = "AU25_t"
    sheet.cell(row=1, column=15).value = "AU26_t"
    sheet.cell(row=1, column=16).value = "AU17_t"
    sheet.cell(row=1, column=17).value = "AU20_t"
    sheet.cell(row=1, column=18).value = "Eyebrow_dist_l"
    sheet.cell(row=1, column=19).value = "Eyebrow_dist_r"
    sheet.cell(row=1, column=20).value = "EAR"
    
    df = pd.read_excel(filename, "CK+_landmarks")
    for i in tqdm(range(df.shape[0])):
        sheet.cell(row=2+i, column=1).value = df.loc[i, "Image path"]
        
        a = (df.loc[i, "Landmark_22_x"], df.loc[i, "Landmark_22_y"])
        b = (df.loc[i, "Landmark_41_x"], df.loc[i, "Landmark_41_y"])
        AU1_l = dist.euclidean(a, b)
        a = (df.loc[i, "Landmark_23_x"], df.loc[i, "Landmark_23_y"])
        b = (df.loc[i, "Landmark_48_x"], df.loc[i, "Landmark_48_y"])
        AU1_r = dist.euclidean(a, b)
        sheet.cell(row=2+i, column=2).value = AU1_l
        sheet.cell(row=2+i, column=3).value = AU1_r
        
        a = (df.loc[i, "Landmark_20_x"], df.loc[i, "Landmark_20_y"])
        b = (df.loc[i, "Landmark_42_x"], df.loc[i, "Landmark_42_y"])
        AU2_l = dist.euclidean(a, b)
        a = (df.loc[i, "Landmark_25_x"], df.loc[i, "Landmark_25_y"])
        b = (df.loc[i, "Landmark_47_x"], df.loc[i, "Landmark_47_y"])
        AU2_r = dist.euclidean(a, b)
        sheet.cell(row=2+i, column=4).value = AU2_l
        sheet.cell(row=2+i, column=5).value = AU2_r
        
        a = (df.loc[i, "Landmark_22_x"], df.loc[i, "Landmark_22_y"])
        b = (df.loc[i, "Landmark_23_x"], df.loc[i, "Landmark_23_y"])
        AU4_t = dist.euclidean(a, b)
        sheet.cell(row=2+i, column=6).value = AU4_t
        
        a = (df.loc[i, "Landmark_22_x"], df.loc[i, "Landmark_22_y"])
        b = (df.loc[i, "Landmark_29_x"], df.loc[i, "Landmark_29_y"])
        AU9_l = dist.euclidean(a, b)
        a = (df.loc[i, "Landmark_23_x"], df.loc[i, "Landmark_23_y"])
        b = (df.loc[i, "Landmark_29_x"], df.loc[i, "Landmark_29_y"])
        AU9_r = dist.euclidean(a, b)
        sheet.cell(row=2+i, column=7).value = AU9_l
        sheet.cell(row=2+i, column=8).value = AU9_r
        
        a = (df.loc[i, "Landmark_49_x"], df.loc[i, "Landmark_49_y"])
        b = (df.loc[i, "Landmark_55_x"], df.loc[i, "Landmark_55_y"])
        AU12_t = dist.euclidean(a, b)
        AU12_l = df.loc[i, "Landmark_49_y"] - df.loc[i, "Landmark_34_y"]
        AU12_r = df.loc[i, "Landmark_55_y"] - df.loc[i, "Landmark_34_y"]
        sheet.cell(row=2+i, column=9).value = AU12_t
        sheet.cell(row=2+i, column=10).value = AU12_l
        sheet.cell(row=2+i, column=11).value = AU12_r
    
        a = (df.loc[i, "Landmark_49_x"], df.loc[i, "Landmark_49_y"])
        b = (df.loc[i, "Landmark_8_x"], df.loc[i, "Landmark_8_y"])
        AU15_l = dist.euclidean(a, b)
        a = (df.loc[i, "Landmark_55_x"], df.loc[i, "Landmark_55_y"])
        b = (df.loc[i, "Landmark_10_x"], df.loc[i, "Landmark_10_y"])
        AU15_r = dist.euclidean(a, b)
        sheet.cell(row=2+i, column=12).value = AU15_l
        sheet.cell(row=2+i, column=13).value = AU15_r
        
        a = (df.loc[i, "Landmark_63_x"], df.loc[i, "Landmark_63_y"])
        b = (df.loc[i, "Landmark_67_x"], df.loc[i, "Landmark_67_y"])
        AU25_t = dist.euclidean(a, b)
        sheet.cell(row=2+i, column=14).value = AU25_t
        
        a = (df.loc[i, "Landmark_34_x"], df.loc[i, "Landmark_34_y"])
        b = (df.loc[i, "Landmark_9_x"], df.loc[i, "Landmark_9_y"])
        AU26_t = dist.euclidean(a, b)
        sheet.cell(row=2+i, column=15).value = AU26_t
        
        a = (df.loc[i, "Landmark_34_x"], df.loc[i, "Landmark_34_y"])
        b = (df.loc[i, "Landmark_9_x"], df.loc[i, "Landmark_9_y"])
        AU17_t = dist.euclidean(a, b)
        sheet.cell(row=2+i, column=16).value = AU17_t
        
        a = (df.loc[i, "Landmark_34_x"], df.loc[i, "Landmark_34_y"])
        b = (df.loc[i, "Landmark_52_x"], df.loc[i, "Landmark_52_y"])
        AU20_t = dist.euclidean(a, b)
        sheet.cell(row=2+i, column=17).value = AU20_t
        
        a = (df.loc[i, "Landmark_20_x"], df.loc[i, "Landmark_20_y"])
        b = (df.loc[i, "Landmark_38_x"], df.loc[i, "Landmark_38_y"])
        eyebrow_dist_l = dist.euclidean(a, b)
        a = (df.loc[i, "Landmark_25_x"], df.loc[i, "Landmark_25_y"])
        b = (df.loc[i, "Landmark_45_x"], df.loc[i, "Landmark_45_y"])
        eyebrow_dist_r = dist.euclidean(a, b)
        sheet.cell(row=2+i, column=18).value = eyebrow_dist_l
        sheet.cell(row=2+i, column=19).value = eyebrow_dist_r
        
        sheet.cell(row=2+i, column=20).value = df.loc[i, "EAR"]

    book.save(filename)

def get_benchmarks_from_CK():
    filename = "CK+_landmarks.xlsx"
    df = pd.read_excel(filename, "geometric_calcs")
    df2 = pd.read_excel(filename, "CK+")
    
    benchmark = []
    AU1_l = []
    AU1_r = []
    AU2_l = []
    AU2_r = []
    AU4_t = []
    AU5_t = []
    AU6_t = []
    AU9_l = []
    AU9_r = []    
    AU12_t = []
    AU12_l = []
    AU12_r = []
    AU15_l = []
    AU15_r = []
    AU25_t = []
    AU26_t = []
    AU17_t = []
    AU20_t = []
    eyebrow_dist = []
    drowsiness = []
    
    #AU1
    for i in tqdm(range(df2.shape[0])):
        if df2.loc[i, "AU1"] == 1:
            epsilon = df.loc[(i*2)+1, "AU1_l"] - df.loc[i*2, "AU1_l"]
            AU1_l.append(epsilon)
            epsilon = df.loc[(i*2)+1, "AU1_r"] - df.loc[i*2, "AU1_r"]
            AU1_r.append(epsilon)
    print(np.percentile(AU1_l, [10, 25, 50, 75]))
    print(np.percentile(AU1_r, [10, 25, 50, 75]))
    benchmark.append(min(np.percentile(AU1_l,25),np.percentile(AU1_r,25)))
    
    #AU2
    for i in tqdm(range(df2.shape[0])):
        if df2.loc[i, "AU2"] == 1:
            epsilon = df.loc[(i*2)+1, "AU2_l"] - df.loc[i*2, "AU2_l"]
            AU2_l.append(epsilon)
            epsilon = df.loc[(i*2)+1, "AU2_r"] - df.loc[i*2, "AU2_r"]
            AU2_r.append(epsilon)
    print(np.mean(AU2_l))
    print(np.percentile(AU2_l, [10, 25, 50, 75]))
    print(np.mean(AU2_r))
    print(np.percentile(AU2_r, [10, 25, 50, 75]))
    benchmark.append(min(np.percentile(AU2_l,25),np.percentile(AU2_r,25)))
    
    #AU4
    for i in tqdm(range(df2.shape[0])):
        if df2.loc[i, "AU4"] == 1:
            epsilon = df.loc[i*2, "AU4_t"] - df.loc[(i*2)+1, "AU4_t"]
            AU4_t.append(epsilon)
    print(np.mean(AU4_t))
    print(np.percentile(AU4_t, [10, 25, 50, 75]))
    benchmark.append(np.percentile(AU4_t,25))
    
    #AU5
    for i in tqdm(range(df2.shape[0])):
        if df2.loc[i, "AU5"] == 1:
            epsilon = df.loc[(i*2)+1, "EAR"] - df.loc[i*2, "EAR"]
            AU5_t.append(epsilon)
    print(np.mean(AU5_t))
    print(np.percentile(AU5_t, [10, 25, 50, 75]))
    benchmark.append(np.percentile(AU5_t,25))
    
    #AU6
    for i in tqdm(range(df2.shape[0])):
        if df2.loc[i, "AU6"] == 1:
            epsilon = df.loc[i*2, "EAR"] - df.loc[(i*2)+1, "EAR"]
            AU6_t.append(epsilon)
    print(np.mean(AU6_t))
    print(np.percentile(AU6_t, [5, 10, 25, 50, 75]))
    benchmark.append(np.percentile(AU6_t,5))
    
    #AU9
    for i in tqdm(range(df2.shape[0])):
        if df2.loc[i, "AU9"] == 1:
            epsilon = df.loc[i*2, "AU9_l"] - df.loc[(i*2)+1, "AU9_l"]
            AU9_l.append(epsilon)
            epsilon = df.loc[i*2, "AU9_r"] - df.loc[(i*2)+1, "AU9_r"]
            AU9_r.append(epsilon)
    print(np.mean(AU9_l))
    print(np.percentile(AU9_l, [15, 25, 50, 75]))
    print(np.mean(AU9_r))
    print(np.percentile(AU9_r, [15, 25, 50, 75]))
    benchmark.append(min(np.percentile(AU9_l,15),np.percentile(AU9_r,15)))
    
    #AU12
    for i in tqdm(range(df2.shape[0])):
        if df2.loc[i, "AU12"] == 1:
            epsilon = df.loc[(i*2)+1, "AU12_t"] - df.loc[i*2, "AU12_t"]
            AU12_t.append(epsilon)
            epsilon = df.loc[i*2, "AU12_l"] - df.loc[(i*2)+1, "AU12_l"]
            AU12_l.append(epsilon)
            epsilon = df.loc[i*2, "AU12_r"] - df.loc[(i*2)+1, "AU12_r"]
            AU12_r.append(epsilon)
    print(np.mean(AU12_t))
    print(np.percentile(AU12_t, [10, 25, 50, 75,90]))
    print(np.mean(AU12_l))
    print(np.percentile(AU12_l, [10, 25, 50, 75,90]))
    print(np.mean(AU12_r))
    print(np.percentile(AU12_r, [10, 25, 50, 75,90]))
    benchmark.append(np.percentile(AU12_t,25))
    benchmark.append(min(np.percentile(AU12_l,25),np.percentile(AU12_r,25)))
    
    #AU15
    for i in tqdm(range(df2.shape[0])):
        if df2.loc[i, "AU15"] == 1:
            epsilon = df.loc[(i*2), "AU15_l"] - df.loc[i*2+1, "AU15_l"]
            AU15_l.append(epsilon)
            epsilon = df.loc[(i*2), "AU15_r"] - df.loc[i*2+1, "AU15_r"]
            AU15_r.append(epsilon)
    print(np.mean(AU15_l))
    print(np.percentile(AU15_l, [10, 25, 50, 75,90]))
    print(np.mean(AU15_r))
    print(np.percentile(AU15_r, [10, 25, 50, 75, 90]))
    benchmark.append(min(np.percentile(AU15_l,25),np.percentile(AU15_r,25)))
    
    #AU25
    for i in tqdm(range(df2.shape[0])):
        if df2.loc[i, "AU25"] == 1:
            epsilon = df.loc[(i*2)+1, "AU25_t"] - df.loc[i*2, "AU25_t"]
            AU25_t.append(epsilon)
    print(np.mean(AU25_t))
    print(np.percentile(AU25_t, [10, 25, 50, 75]))
    benchmark.append(np.percentile(AU25_t,10))
    
    #AU26
    for i in tqdm(range(df2.shape[0])):
        if df2.loc[i, "AU26"] == 1:
            epsilon = df.loc[(i*2)+1, "AU26_t"] - df.loc[i*2, "AU26_t"]
            AU26_t.append(epsilon)
    print(np.mean(AU26_t))
    print(np.percentile(AU26_t, [10, 25, 50, 75]))
    benchmark.append(np.percentile(AU26_t,10))
    
    #AU17
    for i in tqdm(range(df2.shape[0])):
        if df2.loc[i, "AU17"] == 1:
            epsilon = df.loc[(i*2)+1, "AU17_t"] - df.loc[i*2, "AU17_t"]
            AU17_t.append(epsilon)
    print(np.mean(AU17_t))
    print(np.percentile(AU17_t, [10, 25, 50, 75, 90]))
    benchmark.append(np.percentile(AU17_t,25))
    
    #AU20
    for i in tqdm(range(df2.shape[0])):
        if df2.loc[i, "AU20"] == 1:
            epsilon = df.loc[(i*2)+1, "AU20_t"] - df.loc[i*2, "AU20_t"]
            AU20_t.append(epsilon)
    print(np.mean(AU20_t))
    print(np.percentile(AU20_t, [10, 25, 50, 75, 90]))
    benchmark.append(np.percentile(AU20_t,25))
    
    #Normal eyebrow distance
    for i in tqdm(range(df2.shape[0])):
        epsilon = df.loc[i*2, "Eyebrow_dist_l"]
        eyebrow_dist.append(epsilon)
        epsilon = df.loc[i*2, "Eyebrow_dist_l"]
        eyebrow_dist.append(epsilon)
    print(np.mean(eyebrow_dist))
    print(np.percentile(eyebrow_dist, [10, 25, 50, 75, 90]))
    benchmark.append(np.percentile(eyebrow_dist,50))
    
    #Drowsiness - AU43
    for i in tqdm(range(df2.shape[0])):
        if df2.loc[i, "AU43"] == 1:
            epsilon = df.loc[i*2, "EAR"] - df.loc[(i*2)+1, "EAR"]
            drowsiness.append(epsilon)
    print(np.mean(drowsiness))
    print(np.percentile(drowsiness, [10, 25, 50, 75]))
    benchmark.append(np.percentile(drowsiness,50))
    
    benchmark_labels = ["AU1", "AU2", "AU4", "AU5", "AU6", "AU9", "AU12_a", "AU12_b", "AU15", "AU25", "AU26", "AU17", "AU20", "Eyebrow distance", "Drowsiness"]
    with open('benchmark_geometric.csv', 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(["Action Unit Geometric Feature", "Benchmark Value"])
        for i in zip(benchmark_labels, benchmark):
            writer.writerow(i)

def predict_for_val_ars_images():
    filename = "AU_to_Val_Ars.xlsx"
    for tab_name in ("CK+_raw", "KDEF_raw", "Radboud_raw"):
        book = openpyxl.load_workbook(filename)
        sheet = book[tab_name]
        for rownum in tqdm(range(1, len(sheet['A']))):        
            AU_predictions, landmarks = predictor.process_image(sheet.cell(row=rownum+1, column=1).value)
            if not AU_predictions:
                continue
            else:
                for col in range(2, 13):
                    sheet.cell(row = rownum+1, column = col).value = AU_predictions[col - 2]
                landmarks_flatten = landmarks[6].tolist()
                for col in range(1, 69):
                    k = col * 2
                    sheet.cell(row = rownum+1, column = 11+k).value = landmarks_flatten[col-1][0]
                    sheet.cell(row = rownum+1, column = 11+k+1).value = landmarks_flatten[col-1][1]
                sheet.cell(row = rownum+1, column = 149).value = landmarks[7]
        book.save(filename)
    
    for dataset in ("CK+_raw", "KDEF_raw", "Radboud_raw"):
        book = openpyxl.load_workbook(filename)
        #Calculate geometric features
        print("Calculating geometric features...")
        tab_name = dataset[:-3]
        tab_name = "{}geometric_calcs".format(tab_name)
        if tab_name in book.sheetnames:
                sheet = book[tab_name]
                book.remove(sheet)
        sheet = book.create_sheet(title=tab_name)
        sheet.cell(row=1, column=1).value = "Image path"
        sheet.cell(row=1, column=2).value = "AU1_l"
        sheet.cell(row=1, column=3).value = "AU1_r"
        sheet.cell(row=1, column=4).value = "AU2_l"
        sheet.cell(row=1, column=5).value = "AU2_r"
        sheet.cell(row=1, column=6).value = "AU4_t"
        sheet.cell(row=1, column=7).value = "AU9_l"
        sheet.cell(row=1, column=8).value = "AU9_r"
        sheet.cell(row=1, column=9).value = "AU12_t"
        sheet.cell(row=1, column=10).value = "AU12_l"
        sheet.cell(row=1, column=11).value = "AU12_r"
        sheet.cell(row=1, column=12).value = "AU15_l"
        sheet.cell(row=1, column=13).value = "AU15_r"
        sheet.cell(row=1, column=14).value = "AU25_t"
        sheet.cell(row=1, column=15).value = "AU26_t"
        sheet.cell(row=1, column=16).value = "Eyebrow distance"
        sheet.cell(row=1, column=17).value = "EAR"
        
        df = pd.read_excel(filename, dataset)
        for i in tqdm(range(df.shape[0])):
            sheet.cell(row=2+i, column=1).value = df.loc[i, "Image path"]
            
            a = (df.loc[i, "Landmark_22_x"], df.loc[i, "Landmark_22_y"])
            b = (df.loc[i, "Landmark_41_x"], df.loc[i, "Landmark_41_y"])
            AU1_l = dist.euclidean(a, b)
            a = (df.loc[i, "Landmark_23_x"], df.loc[i, "Landmark_23_y"])
            b = (df.loc[i, "Landmark_48_x"], df.loc[i, "Landmark_48_y"])
            AU1_r = dist.euclidean(a, b)
            sheet.cell(row=2+i, column=2).value = AU1_l
            sheet.cell(row=2+i, column=3).value = AU1_r
            
            a = (df.loc[i, "Landmark_20_x"], df.loc[i, "Landmark_20_y"])
            b = (df.loc[i, "Landmark_42_x"], df.loc[i, "Landmark_42_y"])
            AU2_l = dist.euclidean(a, b)
            a = (df.loc[i, "Landmark_25_x"], df.loc[i, "Landmark_25_y"])
            b = (df.loc[i, "Landmark_47_x"], df.loc[i, "Landmark_47_y"])
            AU2_r = dist.euclidean(a, b)
            sheet.cell(row=2+i, column=4).value = AU2_l
            sheet.cell(row=2+i, column=5).value = AU2_r
            
            a = (df.loc[i, "Landmark_22_x"], df.loc[i, "Landmark_22_y"])
            b = (df.loc[i, "Landmark_23_x"], df.loc[i, "Landmark_23_y"])
            AU4_t = dist.euclidean(a, b)
            sheet.cell(row=2+i, column=6).value = AU4_t
            
            a = (df.loc[i, "Landmark_22_x"], df.loc[i, "Landmark_22_y"])
            b = (df.loc[i, "Landmark_29_x"], df.loc[i, "Landmark_29_y"])
            AU9_l = dist.euclidean(a, b)
            a = (df.loc[i, "Landmark_23_x"], df.loc[i, "Landmark_23_y"])
            b = (df.loc[i, "Landmark_29_x"], df.loc[i, "Landmark_29_y"])
            AU9_r = dist.euclidean(a, b)
            sheet.cell(row=2+i, column=7).value = AU9_l
            sheet.cell(row=2+i, column=8).value = AU9_r
            
            a = (df.loc[i, "Landmark_49_x"], df.loc[i, "Landmark_49_y"])
            b = (df.loc[i, "Landmark_55_x"], df.loc[i, "Landmark_55_y"])
            AU12_t = dist.euclidean(a, b)
            AU12_l = df.loc[i, "Landmark_49_y"] - df.loc[i, "Landmark_34_y"]
            AU12_r = df.loc[i, "Landmark_55_y"] - df.loc[i, "Landmark_34_y"]
            sheet.cell(row=2+i, column=9).value = AU12_t
            sheet.cell(row=2+i, column=10).value = AU12_l
            sheet.cell(row=2+i, column=11).value = AU12_r
        
            a = (df.loc[i, "Landmark_49_x"], df.loc[i, "Landmark_49_y"])
            b = (df.loc[i, "Landmark_8_x"], df.loc[i, "Landmark_8_y"])
            AU15_l = dist.euclidean(a, b)
            a = (df.loc[i, "Landmark_55_x"], df.loc[i, "Landmark_55_y"])
            b = (df.loc[i, "Landmark_10_x"], df.loc[i, "Landmark_10_y"])
            AU15_r = dist.euclidean(a, b)
            sheet.cell(row=2+i, column=12).value = AU15_l
            sheet.cell(row=2+i, column=13).value = AU15_r
            
            a = (df.loc[i, "Landmark_63_x"], df.loc[i, "Landmark_63_y"])
            b = (df.loc[i, "Landmark_67_x"], df.loc[i, "Landmark_67_y"])
            AU25_t = dist.euclidean(a, b)
            sheet.cell(row=2+i, column=14).value = AU25_t
            
            a = (df.loc[i, "Landmark_34_x"], df.loc[i, "Landmark_34_y"])
            b = (df.loc[i, "Landmark_9_x"], df.loc[i, "Landmark_9_y"])
            AU26_t = dist.euclidean(a, b)
            sheet.cell(row=2+i, column=15).value = AU26_t
            
            a = (df.loc[i, "Landmark_20_x"], df.loc[i, "Landmark_20_y"])
            b = (df.loc[i, "Landmark_38_x"], df.loc[i, "Landmark_38_y"])
            eyebrow_dist_l = dist.euclidean(a, b)
            a = (df.loc[i, "Landmark_25_x"], df.loc[i, "Landmark_25_y"])
            b = (df.loc[i, "Landmark_45_x"], df.loc[i, "Landmark_45_y"])
            eyebrow_dist_r = dist.euclidean(a, b)
            sheet.cell(row=2+i, column=16).value = min(eyebrow_dist_l, eyebrow_dist_r)
            
            sheet.cell(row=2+i, column=17).value = df.loc[i, "EAR"]
    
        book.save(filename)
        
    #Tweak CK+ data
    book = openpyxl.load_workbook(filename)
    df = pd.read_excel(filename, "CK+_geometric_calcs")
    df2 = pd.read_excel(filename, "CK+_raw")
    df3 = pd.read_csv('benchmark_geometric.csv')
    df3.set_index("Action Unit Geometric Feature", inplace=True)
    raw_data_sheet = book["CK+_raw"]
    if "CK+_raw Copy" in book.sheetnames:
            raw_data_copy_sheet = book["CK+_raw Copy"]
            book.remove(raw_data_copy_sheet)
    book.copy_worksheet(raw_data_sheet)
    sheet = book["CK+_raw Copy"]
    reduction = 3
    min_confidence = 0.5
    
    for i in tqdm(range(int(df2.shape[0]/2))):
        j = i * 2
        
        AU1 = sheet.cell(row=j+2, column=2).value
        AU2a = sheet.cell(row=j+2, column=3).value
        AU2b = sheet.cell(row=j+2, column=4).value
        AU4 = sheet.cell(row=j+2, column=5).value
        AU5 = sheet.cell(row=j+2, column=6).value
        AU6 = sheet.cell(row=j+2, column=7).value
        AU9 = sheet.cell(row=j+2, column=8).value
        AU12 = sheet.cell(row=j+2, column=9).value
        AU15 = sheet.cell(row=j+2, column=10).value
        AU25 = sheet.cell(row=j+2, column=11).value
        AU26 = sheet.cell(row=j+2, column=12).value
        
        if AU1 >= min_confidence:
            if ((df.loc[j+1, "AU1_l"] - df.loc[j, "AU1_l"] < df3.loc["AU1"][0]) and (df.loc[j+1, "AU1_r"] - df.loc[j, "AU1_r"] < df3.loc["AU1"][0])):
                if ((df.loc[j, "AU4_t"] - df.loc[j+1, "AU4_t"]) < df3.loc["AU4"][0]):
                    sheet.cell(row=j+2, column=2).value = AU1/reduction
            
        if AU2a >= min_confidence:
            if ((df.loc[j+1, "AU2_l"] - df.loc[j, "AU2_l"]) < df3.loc["AU2"][0]) and ((df.loc[j+1, "AU2_r"] - df.loc[j, "AU2_r"]) < df3.loc["AU2"][0]):
                if ((df.loc[j, "AU4_t"] - df.loc[j+1, "AU4_t"]) < df3.loc["AU4"][0]):
                    sheet.cell(row=j+2, column=3).value = AU2a/reduction
        
        if AU2a >= min_confidence:
            if ((df.loc[j+1, "AU2_l"] - df.loc[j, "AU2_l"]) < df3.loc["AU2"][0]) and ((df.loc[j+1, "AU2_r"] - df.loc[j, "AU2_r"]) < df3.loc["AU2"][0]):
                if ((df.loc[j, "AU4_t"] - df.loc[j+1, "AU4_t"]) < df3.loc["AU4"][0]):
                    sheet.cell(row=j+2, column=4).value = AU2b/reduction
        
        if AU4 >= min_confidence:
            if ((df.loc[j, "AU4_t"] - df.loc[j+1, "AU4_t"]) < df3.loc["AU4"][0]):
                sheet.cell(row=j+2, column=5).value = AU4/reduction
        
        if AU6 >= min_confidence:
            if ((df.loc[j, "EAR"] - df.loc[j+1, "EAR"]) < df3.loc["AU6"][0]):
                sheet.cell(row=j+2, column=7).value = AU6/reduction
        
        if AU9 >= min_confidence:
            if ((df.loc[j, "AU9_l"] - df.loc[j+1, "AU9_l"]) < df3.loc["AU9"][0]):
                sheet.cell(row=j+2, column=8).value = AU9/reduction
            elif ((df.loc[j, "AU9_r"] - df.loc[j+1, "AU9_r"]) < df3.loc["AU9"][0]):
                sheet.cell(row=j+2, column=8).value = AU9/reduction
        
        if AU15 >= min_confidence:
            if ((df.loc[j, "AU15_l"] - df.loc[j+1, "AU15_l"]) < df3.loc["AU15"][0]):
                sheet.cell(row=j+2, column=10).value = AU15/reduction
            elif ((df.loc[j, "AU15_r"] - df.loc[j+1, "AU15_r"]) < df3.loc["AU15"][0]):
                sheet.cell(row=j+2, column=10).value = AU15/reduction
        
        if AU25 >= min_confidence:
            if ((df.loc[j+1, "AU25_t"] - df.loc[j, "AU25_t"]) < df3.loc["AU25"][0]):
                sheet.cell(row=j+2, column=11).value = AU25/reduction
        
        if AU26 >= min_confidence:
            if ((df.loc[j+1, "AU26_t"] - df.loc[j, "AU26_t"]) < df3.loc["AU26"][0]):
                sheet.cell(row=j+2, column=12).value = AU26/reduction
    
    book.save(filename)
    
    #Tweak KDEF data
    book = openpyxl.load_workbook(filename)
    df = pd.read_excel(filename, "KDEF_geometric_calcs")
    df.set_index("Image path", inplace=True)
    df2 = pd.read_excel(filename, "KDEF_raw")
    df3 = pd.read_csv('benchmark_geometric.csv')
    df3.set_index("Action Unit Geometric Feature", inplace=True)
    raw_data_sheet = book["KDEF_raw"]
    if "KDEF_raw Copy" in book.sheetnames:
            raw_data_copy_sheet = book["KDEF_raw Copy"]
            book.remove(raw_data_copy_sheet)
    book.copy_worksheet(raw_data_sheet)
    sheet = book["KDEF_raw Copy"]
    reduction = 3
    min_confidence = 0.5
    
    for i in tqdm(range(df2.shape[0])):
        j = df2.loc[i, "Image path"][:-7] + "NES.JPG"
        image = df2.loc[i, "Image path"]
        neutral = df.loc[j]

        AU1 = sheet.cell(row=i+2, column=2).value
        AU2a = sheet.cell(row=i+2, column=3).value
        AU2b = sheet.cell(row=i+2, column=4).value
        AU4 = sheet.cell(row=i+2, column=5).value
        AU5 = sheet.cell(row=i+2, column=6).value
        AU6 = sheet.cell(row=i+2, column=7).value
        AU9 = sheet.cell(row=i+2, column=8).value
        AU12 = sheet.cell(row=i+2, column=9).value
        AU15 = sheet.cell(row=i+2, column=10).value
        AU25 = sheet.cell(row=i+2, column=11).value
        AU26 = sheet.cell(row=i+2, column=12).value
        
        if AU1 >= min_confidence:
            if ((df.loc[image, "AU1_l"] - neutral["AU1_l"] < df3.loc["AU1"][0]) and (df.loc[image, "AU1_r"] - neutral["AU1_r"] < df3.loc["AU1"][0])):
                if ((neutral["AU4_t"] - df.loc[image, "AU4_t"]) < df3.loc["AU4"][0]):
                    sheet.cell(row=i+2, column=2).value = AU1/reduction
        
        if AU2a >= min_confidence:
            if ((df.loc[image, "AU2_l"] - neutral["AU2_l"]) < df3.loc["AU2"][0]) and ((df.loc[image, "AU2_r"] - neutral["AU2_r"]) < df3.loc["AU2"][0]):
                if ((neutral["AU4_t"] - df.loc[image, "AU4_t"]) < df3.loc["AU4"][0]):
                    sheet.cell(row=i+2, column=3).value = AU2a/reduction
        
        if AU2a >= min_confidence:
            if ((df.loc[image, "AU2_l"] - neutral["AU2_l"]) < df3.loc["AU2"][0]) and ((df.loc[image, "AU2_r"] - neutral["AU2_r"]) < df3.loc["AU2"][0]):
                if ((neutral["AU4_t"] - df.loc[image, "AU4_t"]) < df3.loc["AU4"][0]):
                    sheet.cell(row=i+2, column=4).value = AU2b/reduction
        
        if AU4 >= min_confidence:
            if ((neutral["AU4_t"] - df.loc[image, "AU4_t"]) < df3.loc["AU4"][0]):
                sheet.cell(row=i+2, column=5).value = AU4/reduction
        
        if AU6 >= min_confidence:
            if ((neutral["EAR"] - df.loc[image, "EAR"]) < df3.loc["AU6"][0]):
                sheet.cell(row=i+2, column=7).value = AU6/reduction
        
        if AU9 >= min_confidence:
            if ((neutral["AU9_l"] - df.loc[image, "AU9_l"]) < df3.loc["AU9"][0]):
                sheet.cell(row=i+2, column=8).value = AU9/reduction
            elif ((neutral["AU9_r"] - df.loc[image, "AU9_r"]) < df3.loc["AU9"][0]):
                sheet.cell(row=i+2, column=8).value = AU9/reduction
        
        if AU15 >= min_confidence:
            if ((neutral["AU15_l"] - df.loc[image, "AU15_l"]) < df3.loc["AU15"][0]):
                sheet.cell(row=i+2, column=10).value = AU15/reduction
            elif ((neutral["AU15_r"] - df.loc[image, "AU15_r"]) < df3.loc["AU15"][0]):
                sheet.cell(row=i+2, column=10).value = AU15/reduction
        
        if AU25 >= min_confidence:
            if ((df.loc[image, "AU25_t"] - neutral["AU25_t"]) < df3.loc["AU25"][0]):
                sheet.cell(row=i+2, column=11).value = AU25/reduction
        
        if AU26 >= min_confidence:
            if ((df.loc[image, "AU26_t"] - neutral["AU26_t"]) < df3.loc["AU26"][0]):
                sheet.cell(row=i+2, column=12).value = AU26/reduction
    
    book.save(filename)
    
    #Tweak Radboud data
    book = openpyxl.load_workbook(filename)
    df = pd.read_excel(filename, "Radboud_geometric_calcs")
    df.set_index("Image path", inplace=True)
    df2 = pd.read_excel(filename, "Radboud_raw")
    df3 = pd.read_csv('benchmark_geometric.csv')
    df3.set_index("Action Unit Geometric Feature", inplace=True)
    raw_data_sheet = book["Radboud_raw"]
    if "Radboud_raw Copy" in book.sheetnames:
            raw_data_copy_sheet = book["Radboud_raw Copy"]
            book.remove(raw_data_copy_sheet)
    book.copy_worksheet(raw_data_sheet)
    sheet = book["Radboud_raw Copy"]
    reduction = 3
    min_confidence = 0.5
    
    for i in tqdm(range(df2.shape[0])):
        subject = df2.loc[i, "Image path"][50:60]
        image = df2.loc[i, "Image path"]
        subject_images = df.filter(like=subject, axis=0)
        neutral = subject_images.filter(like="neutral_frontal", axis=0)
        neutral = neutral.iloc[0]
        
        AU1 = sheet.cell(row=i+2, column=2).value
        AU2a = sheet.cell(row=i+2, column=3).value
        AU2b = sheet.cell(row=i+2, column=4).value
        AU4 = sheet.cell(row=i+2, column=5).value
        AU5 = sheet.cell(row=i+2, column=6).value
        AU6 = sheet.cell(row=i+2, column=7).value
        AU9 = sheet.cell(row=i+2, column=8).value
        AU12 = sheet.cell(row=i+2, column=9).value
        AU15 = sheet.cell(row=i+2, column=10).value
        AU25 = sheet.cell(row=i+2, column=11).value
        AU26 = sheet.cell(row=i+2, column=12).value
        
        if AU1 >= min_confidence:
            if ((df.loc[image, "AU1_l"] - neutral["AU1_l"] < df3.loc["AU1"][0]) and (df.loc[image, "AU1_r"] - neutral["AU1_r"] < df3.loc["AU1"][0])):
                if ((neutral["AU4_t"] - df.loc[image, "AU4_t"]) < df3.loc["AU4"][0]):
                    sheet.cell(row=i+2, column=2).value = AU1/reduction
        
        if AU2a >= min_confidence:
            if ((df.loc[image, "AU2_l"] - neutral["AU2_l"]) < df3.loc["AU2"][0]) and ((df.loc[image, "AU2_r"] - neutral["AU2_r"]) < df3.loc["AU2"][0]):
                if ((neutral["AU4_t"] - df.loc[image, "AU4_t"]) < df3.loc["AU4"][0]):
                    sheet.cell(row=i+2, column=3).value = AU2a/reduction
        
        if AU2a >= min_confidence:
            if ((df.loc[image, "AU2_l"] - neutral["AU2_l"]) < df3.loc["AU2"][0]) and ((df.loc[image, "AU2_r"] - neutral["AU2_r"]) < df3.loc["AU2"][0]):
                if ((neutral["AU4_t"] - df.loc[image, "AU4_t"]) < df3.loc["AU4"][0]):
                    sheet.cell(row=i+2, column=4).value = AU2b/reduction
        
        if AU4 >= min_confidence:
            if ((neutral["AU4_t"] - df.loc[image, "AU4_t"]) < df3.loc["AU4"][0]):
                sheet.cell(row=i+2, column=5).value = AU4/reduction
        
        if AU6 >= min_confidence:
            if ((neutral["EAR"] - df.loc[image, "EAR"]) < df3.loc["AU6"][0]):
                sheet.cell(row=i+2, column=7).value = AU6/reduction
        
        if AU9 >= min_confidence:
            if ((neutral["AU9_l"] - df.loc[image, "AU9_l"]) < df3.loc["AU9"][0]):
                sheet.cell(row=i+2, column=8).value = AU9/reduction
            elif ((neutral["AU9_r"] - df.loc[image, "AU9_r"]) < df3.loc["AU9"][0]):
                sheet.cell(row=i+2, column=8).value = AU9/reduction
        
        if AU15 >= min_confidence:
            if ((neutral["AU15_l"] - df.loc[image, "AU15_l"]) < df3.loc["AU15"][0]):
                sheet.cell(row=i+2, column=10).value = AU15/reduction
            elif ((neutral["AU15_r"] - df.loc[image, "AU15_r"]) < df3.loc["AU15"][0]):
                sheet.cell(row=i+2, column=10).value = AU15/reduction
        
        if AU25 >= min_confidence:
            if ((df.loc[image, "AU25_t"] - neutral["AU25_t"]) < df3.loc["AU25"][0]):
                sheet.cell(row=i+2, column=11).value = AU25/reduction
        
        if AU26 >= min_confidence:
            if ((df.loc[image, "AU26_t"] - neutral["AU26_t"]) < df3.loc["AU26"][0]):
                sheet.cell(row=i+2, column=12).value = AU26/reduction
    
    book.save(filename)
        
    #Final processed output with valence and arousal
    for dataset in ("CK+_raw Copy", "KDEF_raw Copy", "Radboud_raw Copy"):
        book = openpyxl.load_workbook(filename)
        df = pd.read_excel(filename, dataset)
        geometric_tab = dataset[:-8]
        geometric_tab = "{}geometric_calcs".format(geometric_tab)
        df2 = pd.read_excel(filename, geometric_tab)
        df3 = pd.read_csv('benchmark_geometric.csv')
        df3.set_index("Action Unit Geometric Feature", inplace=True)
        #Calculate geometric features
        print("Processing data...")
        tab_name = dataset[:-8]
        tab_name = "{}processed".format(tab_name)
        if tab_name in book.sheetnames:
                sheet = book[tab_name]
                book.remove(sheet)
        sheet = book.create_sheet(title=tab_name)
        
        sheet.cell(row=1, column=1).value = "Image path"
        sheet.cell(row=1, column=2).value = "AU1"
        sheet.cell(row=1, column=3).value = "AU2"
        sheet.cell(row=1, column=4).value = "AU4"
        sheet.cell(row=1, column=5).value = "AU5"
        sheet.cell(row=1, column=6).value = "AU6"
        sheet.cell(row=1, column=7).value = "AU9"
        sheet.cell(row=1, column=8).value = "AU12"
        sheet.cell(row=1, column=9).value = "AU15"
        sheet.cell(row=1, column=10).value = "AU25"
        sheet.cell(row=1, column=11).value = "AU26"
        sheet.cell(row=1, column=12).value = "Emotion"
        sheet.cell(row=1, column=13).value = "Valence"
        sheet.cell(row=1, column=14).value = "Arousal"
        
        for i in tqdm(range(df.shape[0])):
            sheet.cell(row=i+2, column=1).value = df.loc[i, "Image path"]
            if df2.loc[i, "Eyebrow distance"] > df3.loc["Eyebrow distance"][0]:
                AU2_col = 2 
            else:
                AU2_col = 3
            for j in range(1, 11):
                if j == 1:
                    sheet.cell(row=i+2, column=j+1).value = df.iloc[i, j]
                elif j == 2:
                    sheet.cell(row=i+2, column=j+1).value = df.iloc[i, AU2_col]
                else:
                    sheet.cell(row=i+2, column=j+1).value = df.iloc[i, j+1]
        
        if dataset == "CK+_raw Copy":
            path = os.getcwd() + "\data\CKPlusDataset\Emotion"
            for folder in os.listdir(path):
                folder_path = path + "\\" + folder
    
                for image_folder in os.listdir(folder_path):
                    image_folder_path = folder_path + "\\" + image_folder
    
                    for textfile in glob.glob(os.path.join(image_folder_path, '*.txt')):
                        with open(textfile) as f:
                            content = f.readlines()
                            for rownum in tqdm(range(df.shape[0])):
                                if (df.loc[rownum,"Image path"][61:-4] == textfile[50:-12]):
                                    emotion_number = int(float(content[0][3:16]))
                                    sheet.cell(row=rownum+2, column=12).value = emotion_number
                                    sheet.cell(row=rownum+2, column=13).value = DATASET.mapping_emotion_number_to_valence[emotion_number]
                                    sheet.cell(row=rownum+2, column=14).value = DATASET.mapping_emotion_number_to_arousal[emotion_number]
                                    break
        
        if dataset == "KDEF_raw Copy":
            emotion_to_number = {"AF":4, "AN":1, "DI":3, "HA":5, "NE":0, "SA":6, "SU":7}
            for rownum in tqdm(range(df.shape[0])):
                emotion_label = df.loc[rownum,"Image path"][-7:-5]
                emotion_number = emotion_to_number[emotion_label]
                sheet.cell(row=rownum+2, column=12).value = emotion_number
                sheet.cell(row=rownum+2, column=13).value = DATASET.mapping_emotion_number_to_valence[emotion_number]
                sheet.cell(row=rownum+2, column=14).value = DATASET.mapping_emotion_number_to_arousal[emotion_number] 

        if dataset == "Radboud_raw Copy":
            df4 = pd.read_excel('data/Radboud/radboud_valence.xlsx')
            df4.set_index("Image", inplace=True)
            emotion_to_number = {"angry":1, "contemptuous":2, "disgusted":3, "fearful":4, "happy":5, "neutral":0, "sad":6, "surprised":7}
            for rownum in tqdm(range(df.shape[0])):
                emotion_label = None
                for label in ("angry_", "contemptuous_", "disgusted_", "fearful_", "happy_", "neutral_", "sad_", "surprised_"):
                    if label in df.loc[rownum,"Image path"]:
                        emotion_label = label[:-1]
                        break
                emotion_number = emotion_to_number[emotion_label]
                sheet.cell(row=rownum+2, column=12).value = emotion_number
                sheet.cell(row=rownum+2, column=14).value = DATASET.mapping_emotion_number_to_arousal[emotion_number]
                image_path_val = df.loc[rownum,"Image path"][50:-4]
                sheet.cell(row=rownum+2, column=13).value = df4.loc[image_path_val,"Valence scaled"]
        
        book.save(filename)
    
    #Adding neutral values to CK+ labels
    book = openpyxl.load_workbook(filename)
    tab = "CK+_processed"
    df = pd.read_excel(filename, tab)
    sheet = book[tab]
    for i in range(df.shape[0]):
        if df.loc[i, "Image path"][-7:] == "001.png":
            sheet.cell(row=i+2, column=12).value = 0
            sheet.cell(row=i+2, column=13).value = 0
            sheet.cell(row=i+2, column=14).value = 0
    book.save(filename)
    
    #Making all neutral images to have 0 for AUs
    filename = "AU_to_Val_Ars - Copy.xlsx"
    for dataset in ("CK+_processed", "KDEF_processed", "Radboud_processed"):
        book = openpyxl.load_workbook(filename)
        df = pd.read_excel(filename, dataset)
        sheet = book[dataset]
        for i in range(df.shape[0]):
            if df.loc[i, "Emotion"] == 0:
                for j in range(10):
                    sheet.cell(row=i+2, column=j+2).value = 0
        book.save(filename)

if __name__ == "__main__":
    predictor = Photo_Predictor()