import matplotlib.pyplot as plt
plt.switch_backend('agg')
import numpy as np
import cv2
import os
import sys
import dlib
from imutils import face_utils
from head_pose_estimator import Headpose_Estimator
from scipy.spatial import distance as dist
from keras.models import load_model
import openpyxl
from tqdm import tqdm
import pandas as pd
from sklearn.externals import joblib
from sklearn.preprocessing import MinMaxScaler
from parameters import DATASET, VIDEO_PREDICTOR, OTHER
from matplotlib.ticker import StrMethodFormatter
np.random.seed(OTHER.random_state)
#%matplotlib inline

class Video_Predictor:

    def __init__(self):

        #Load all models
        self.net = cv2.dnn.readNetFromCaffe(DATASET.config_file, DATASET.model_file)
        self.face_landmarks_predictor = dlib.shape_predictor(VIDEO_PREDICTOR.face_landmarks_file)
        self.secondary_face_detector = dlib.get_frontal_face_detector()
        self.model_AU1 = load_model('models/AU1/AU1.h5')
        self.model_AU2a = load_model('models/AU2/AU2a.h5')
        self.model_AU2b = load_model('models/AU2/AU2b.h5')
        self.model_AU4 = load_model('models/AU4/AU4.h5')
        self.model_AU5 = load_model('models/AU5/AU5.h5')
        self.model_AU6 = load_model('models/AU6/AU6.h5')
        self.model_AU9 = load_model('models/AU9/AU9.h5')
        self.model_AU12 = load_model('models/AU12/AU12.h5')
        self.model_AU15 = load_model('models/AU15/AU15.h5')
        self.model_AU25 = load_model('models/AU25/AU25.h5')
        self.model_AU26 = load_model('models/AU26/AU26.h5')
        (self.lStart, self.lEnd) = face_utils.FACIAL_LANDMARKS_IDXS["left_eye"]
        (self.rStart, self.rEnd) = face_utils.FACIAL_LANDMARKS_IDXS["right_eye"]
        self.hpe = Headpose_Estimator()

    def aggregate_AU_predictions(self, only_face_image):
        AU_predictions_per_frame = []

        image_grayscale = cv2.cvtColor(only_face_image, cv2.COLOR_RGB2GRAY)

        # resize to normalize data size
        whole_face = cv2.resize(image_grayscale, (DATASET.face_horizontal_size, DATASET.face_vertical_size))

        # Obtain specific faces including flipped images
        upper_face = whole_face[0:DATASET.upper_face_vertical_size, ].copy()
        upper_face_flip = cv2.flip(upper_face, 1)
        lower_face = whole_face[115:(115+DATASET.lower_face_vertical_size), ].copy()
        lower_face_flip = cv2.flip(lower_face, 1)
        whole_face = cv2.resize(whole_face, (DATASET.face_horizontal_size, DATASET.face_vertical_size))
        whole_face_flip = cv2.flip(whole_face, 1)

        #Reshape all faces for model input
        upper_face = np.array(upper_face).reshape(-1, DATASET.upper_face_vertical_size, DATASET.face_horizontal_size, 1)
        upper_face_flip = np.array(upper_face_flip).reshape(-1, DATASET.upper_face_vertical_size, DATASET.face_horizontal_size, 1)
        whole_face = np.array(whole_face).reshape(-1, DATASET.face_vertical_size, DATASET.face_horizontal_size, 1)
        whole_face_flip = np.array(whole_face_flip).reshape(-1, DATASET.face_vertical_size, DATASET.face_horizontal_size, 1)
        lower_face = np.array(lower_face).reshape(-1, DATASET.lower_face_vertical_size, DATASET.face_horizontal_size, 1)
        lower_face_flip = np.array(lower_face_flip).reshape(-1, DATASET.lower_face_vertical_size, DATASET.face_horizontal_size, 1)

        # Scale all images
        upper_face = upper_face/255
        upper_face_flip = upper_face_flip/255
        whole_face = whole_face/255
        whole_face_flip = whole_face_flip/255
        lower_face = lower_face/255
        lower_face_flip = lower_face_flip/255

        #Predict for all faces
        AU1a = self.model_AU1.predict(upper_face)
        AU1b = self.model_AU1.predict(upper_face_flip)
        AU1 = [[AU1a[0][0]*0.5 + AU1b[0][0]*0.5, AU1a[0][1]*0.5 + AU1b[0][1]*0.5]]
        AU2a_1 = self.model_AU2a.predict(upper_face)
        AU2a_2 = self.model_AU2a.predict(upper_face_flip)
        AU2a = [[AU2a_1[0][0]*0.5 + AU2a_2[0][0]*0.5, AU2a_1[0][1]*0.5 + AU2a_2[0][1]*0.5]]
        AU2b_1 = self.model_AU2b.predict(upper_face)
        AU2b_2 = self.model_AU2b.predict(upper_face_flip)
        AU2b = [[AU2b_1[0][0]*0.5 + AU2b_2[0][0]*0.5, AU2b_1[0][1]*0.5 + AU2b_2[0][1]*0.5]]
        AU4a = self.model_AU4.predict(upper_face)
        AU4b = self.model_AU4.predict(upper_face_flip)
        AU4 = [[AU4a[0][0]*0.5 + AU4b[0][0]*0.5, AU4a[0][1]*0.5 + AU4b[0][1]*0.5]]
        AU5a = self.model_AU5.predict(upper_face)
        AU5b = self.model_AU5.predict(upper_face_flip)
        AU5 = [[AU5a[0][0]*0.5 + AU5b[0][0]*0.5, AU5a[0][1]*0.5 + AU5b[0][1]*0.5]]
        AU6a = self.model_AU6.predict(whole_face)
        AU6b = self.model_AU6.predict(whole_face_flip)
        AU6 = [[AU6a[0][0]*0.5 + AU6b[0][0]*0.5, AU6a[0][1]*0.5 + AU6b[0][1]*0.5]]
        AU9a = self.model_AU9.predict(whole_face)
        AU9b = self.model_AU9.predict(whole_face_flip)
        AU9 = [[AU9a[0][0]*0.5 + AU9b[0][0]*0.5, AU9a[0][1]*0.5 + AU9b[0][1]*0.5]]
        AU12a = self.model_AU12.predict(lower_face)
        AU12b = self.model_AU12.predict(lower_face_flip)
        AU12 = [[AU12a[0][0]*0.5 + AU12b[0][0]*0.5, AU12a[0][1]*0.5 + AU12b[0][1]*0.5]]
        AU15a = self.model_AU15.predict(lower_face)
        AU15b = self.model_AU15.predict(lower_face_flip)
        AU15 = [[AU15a[0][0]*0.5 + AU15b[0][0]*0.5, AU15a[0][1]*0.5 + AU15b[0][1]*0.5]]
        AU25a = self.model_AU25.predict(lower_face)
        AU25b = self.model_AU25.predict(lower_face_flip)
        AU25 = [[AU25a[0][0]*0.5 + AU25b[0][0]*0.5, AU25a[0][1]*0.5 + AU25b[0][1]*0.5]]
        AU26a = self.model_AU26.predict(lower_face)
        AU26b = self.model_AU26.predict(lower_face_flip)
        AU26 = [[AU26a[0][0]*0.5 + AU26b[0][0]*0.5, AU26a[0][1]*0.5 + AU26b[0][1]*0.5]]

        #Append all predictions into list of predictions for the frame
        AU_predictions_per_frame.append(AU1[0][1])
        AU_predictions_per_frame.append(AU2a[0][1])
        AU_predictions_per_frame.append(AU2b[0][1])
        AU_predictions_per_frame.append(AU4[0][1])
        AU_predictions_per_frame.append(AU5[0][1])
        AU_predictions_per_frame.append(AU6[0][1])
        AU_predictions_per_frame.append(AU9[0][1])
        AU_predictions_per_frame.append(AU12[0][1])
        AU_predictions_per_frame.append(AU15[0][1])
        AU_predictions_per_frame.append(AU25[0][1])
        AU_predictions_per_frame.append(AU26[0][1])

        return AU_predictions_per_frame

    def rotate(self, image, angle, center=None, scale=1.0):
        # grab the dimensions of the image
        (h, w) = image.shape[:2]

        # if the center is None, initialize it as the center of
        # the image
        if center is None:
            center = (w // 2, h // 2)

        # perform the rotation
        M = cv2.getRotationMatrix2D(center, angle, scale)
        rotated = cv2.warpAffine(image, M, (w, h))

        # return the rotated image
        return rotated

    def process_video(self, filename):
        # Playing video from file:
        cap = cv2.VideoCapture(filename)
        # Check if camera opened successfully
        if (cap.isOpened()== False):
            print("Error opening video stream or file")
        frame_count = 0
        predictions = []

        # Read until video is completed
        while(cap.isOpened()):
          # Capture frame-by-frame
          ret, image = cap.read()
          only_face_image = image
          AU_predictions_per_frame = []
          landmarks_per_frame = []

          if ret == True:
            frame_count += 1
            if frame_count%3 == 0 or frame_count%3 == 2:
                continue
            print(frame_count)
            # detect faces
            (h, w) = image.shape[:2]
            #Rotate the face so it is aligned horizontally i.e. Roll is at 0 degrees
            angle, shape = self.get_rotation_angle(image)
            non_rotated_image = image.copy()
            image = self.rotate(image, angle)
            blob = cv2.dnn.blobFromImage(cv2.resize(image, (300,300)), 1.0, (300, 300), (104.0, 177.0, 123.0))
            self.net.setInput(blob)
            detections = self.net.forward()
            # loop over the detections and pick highest confidence picture
            detection_index = 0
            max_confidence = 0
            for i in range(0, detections.shape[2]):
                confidence = detections[0, 0, i, 2]
                if max_confidence < confidence:
                    max_confidence = confidence
                    detection_index = i

            i = detection_index
    		# extract the confidence (i.e., probability) associated with the
    		# prediction
            confidence = detections[0, 0, i, 2]
    		# filter out weak detections by ensuring the `confidence` is
    		# greater than the minimum confidence
            if ((confidence <= DATASET.min_confidence) or (detections.shape[2] == 0) or (shape is None)):
                predictions.append([AU_predictions_per_frame, landmarks_per_frame])
                continue
    		# compute the (x, y)-coordinates of the bounding box for the
    		# object
            box = detections[0, 0, i, 3:7] * np.array([w, h, w, h])
            (startX, startY, endX, endY) = box.astype("int")
            only_face_image = image[startY-2:endY+2, startX-2:endX+2].copy()

            #Send aligned face for landmark detection
            landmarks_per_frame = self.landmark_detection(non_rotated_image, shape)
            #Aligned face image to be detected again for face and sent for AU detection
            AU_predictions_per_frame = self.aggregate_AU_predictions(only_face_image)
            predictions.append([AU_predictions_per_frame, landmarks_per_frame])

          else:
            break
        cap.release()
        cv2.destroyAllWindows()
        return predictions

    def landmark_detection(self, image, shape):
        landmarks_per_frame = []
        #Head pose estimations
        angles = self.hpe.process_image(image, shape)
        pitch_min = angles[0]
        pitch_avg = angles[0]
        yaw_min = angles[1]
        yaw_avg = angles[1]
        roll_min = angles[2]
        roll_avg = angles[2]
        landmarks_per_frame.append(pitch_min)
        landmarks_per_frame.append(pitch_avg)
        landmarks_per_frame.append(yaw_min)
        landmarks_per_frame.append(yaw_avg)
        landmarks_per_frame.append(roll_min)
        landmarks_per_frame.append(roll_avg)

        #Facial landmarks
        shape = face_utils.shape_to_np(shape)
        landmarks_per_frame.append(shape)
        #Detect drowsiness
        drowsiness_per_frame = self.drowsiness_detection(shape)
        landmarks_per_frame.append(drowsiness_per_frame)

        return landmarks_per_frame

    def drowsiness_detection(self, shape):
        # extract the left and right eye coordinates, then use the
        # coordinates to compute the eye aspect ratio for both eyes
        leftEye = shape[self.lStart:self.lEnd]
        rightEye = shape[self.rStart:self.rEnd]
        leftEAR = self.eye_aspect_ratio(leftEye)
        rightEAR = self.eye_aspect_ratio(rightEye)

        # average the eye aspect ratio together for both eyes
        ear = (leftEAR + rightEAR) / 2.0

        return ear

    def eye_aspect_ratio(self, eye):
    	# compute the euclidean distances between the two sets of
    	# vertical eye landmarks (x, y)-coordinates
    	A = dist.euclidean(eye[1], eye[5])
    	B = dist.euclidean(eye[2], eye[4])

    	# compute the euclidean distance between the horizontal
    	# eye landmark (x, y)-coordinates
    	C = dist.euclidean(eye[0], eye[3])

    	# compute the eye aspect ratio
    	ear = (A + B) / (2.0 * C)

    	# return the eye aspect ratio
    	return ear

    def get_rotation_angle(self, image):
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        rects = self.secondary_face_detector(gray, 1)
        angle = 0
        landmarks = None
        if len(rects) > 0:
            # convert the landmark (x, y)-coordinates to a NumPy array
            landmarks = self.face_landmarks_predictor(gray, rects[0])
            shape = face_utils.shape_to_np(landmarks)

            if (len(shape)==68):
    			# extract the left and right eye (x, y)-coordinates
                (lStart, lEnd) = face_utils.FACIAL_LANDMARKS_68_IDXS["left_eye"]
                (rStart, rEnd) = face_utils.FACIAL_LANDMARKS_68_IDXS["right_eye"]
            else:
                (lStart, lEnd) = face_utils.FACIAL_LANDMARKS_5_IDXS["left_eye"]
                (rStart, rEnd) = face_utils.FACIAL_LANDMARKS_5_IDXS["right_eye"]

            leftEyePts = shape[lStart:lEnd]
            rightEyePts = shape[rStart:rEnd]

    		# compute the center of mass for each eye
            leftEyeCenter = leftEyePts.mean(axis=0).astype("int")
            rightEyeCenter = rightEyePts.mean(axis=0).astype("int")

    		# compute the angle between the eye centroids
            dY = rightEyeCenter[1] - leftEyeCenter[1]
            dX = rightEyeCenter[0] - leftEyeCenter[0]
            angle = np.degrees(np.arctan2(dY, dX)) - 180
        return angle, landmarks

def save_to_excel(predictions, filename):
    filename, ext = os.path.splitext(filename)
    filename = filename + ".xlsx"
    if filename in os.listdir(os.getcwd()):
        book = openpyxl.load_workbook(filename)
        if "raw_data" in book.sheetnames:
            sheet = book["raw_data"]
            book.remove_sheet(sheet)
    else:
        book = openpyxl.Workbook()

    sheet = book.create_sheet(title="raw_data")
    sheet.cell(row=1, column=1).value = "Frame"
    sheet.cell(row=1, column=2).value = "AU1"
    sheet.cell(row=1, column=3).value = "AU2a"
    sheet.cell(row=1, column=4).value = "AU2b"
    sheet.cell(row=1, column=5).value = "AU4"
    sheet.cell(row=1, column=6).value = "AU5"
    sheet.cell(row=1, column=7).value = "AU6"
    sheet.cell(row=1, column=8).value = "AU9"
    sheet.cell(row=1, column=9).value = "AU12"
    sheet.cell(row=1, column=10).value = "AU15"
    sheet.cell(row=1, column=11).value = "AU25"
    sheet.cell(row=1, column=12).value = "AU26"
    sheet.cell(row=1, column=13).value = "Pitch_Min"
    sheet.cell(row=1, column=14).value = "Pitch_Avg"
    sheet.cell(row=1, column=15).value = "Yaw_Min"
    sheet.cell(row=1, column=16).value = "Yaw_Avg"
    sheet.cell(row=1, column=17).value = "Roll_Min"
    sheet.cell(row=1, column=18).value = "Roll_Avg"
    for i in range(1, 69):
        j = i * 2
        sheet.cell(row=1, column=(18+(j-1))).value = "Landmark_{}_x".format(i)
        sheet.cell(row=1, column=(18+j)).value = "Landmark_{}_y".format(i)
    sheet.cell(row=1, column=155).value = "EAR"

    for j in tqdm(range(len(predictions))):
        sheet.cell(row = j + 2, column = 1).value = j + 1
        # Adding AU predictions into excel if not empty
        if not predictions[j][0]:
            for col in range(2, 13):
                sheet.cell(row = j + 2, column = col).value = 0
        else:
            for col in range(2, 13):
                sheet.cell(row = j + 2, column = col).value = predictions[j][0][col - 2]

        # Adding landmarks estimations into excel if not empty
        if not predictions[j][1]:
            for col in range(13, 88):
                sheet.cell(row = j + 2, column = col).value = 0
        else:
            for col in range(13, 19):
                sheet.cell(row = j + 2, column = col).value = predictions[j][1][col - 13]
            landmarks_flatten = predictions[j][1][6].tolist()
            for col in range(1, 69):
                k = col * 2
                sheet.cell(row = j + 2, column = (18+k-1)).value = landmarks_flatten[col-1][0]
                sheet.cell(row = j + 2, column = (18+k)).value = landmarks_flatten[col-1][1]
            sheet.cell(row = j + 2, column = 155).value = predictions[j][1][7]

    book.save(filename)
    process_raw_data(filename)
    create_valence_and_arousal_charts(filename)

def process_raw_data(filename):
    book = openpyxl.load_workbook(filename)

    #Calculate benchmark figures
    print("Calculating geometric features...")
    if "geometric_calcs" in book.sheetnames:
            sheet = book["geometric_calcs"]
            book.remove(sheet)
    sheet = book.create_sheet(title="geometric_calcs")
    sheet.cell(row=1, column=1).value = "Frame"
    sheet.cell(row=1, column=2).value = "AU1_l"
    sheet.cell(row=1, column=3).value = "AU1_r"
    sheet.cell(row=1, column=4).value = "AU2_l"
    sheet.cell(row=1, column=5).value = "AU2_r"
    sheet.cell(row=1, column=6).value = "AU4_t"
    sheet.cell(row=1, column=7).value = "AU5_t"
    sheet.cell(row=1, column=8).value = "AU6_t"
    sheet.cell(row=1, column=9).value = "AU9_l"
    sheet.cell(row=1, column=10).value = "AU9_r"
    sheet.cell(row=1, column=11).value = "AU12_t"
    sheet.cell(row=1, column=12).value = "AU12_l"
    sheet.cell(row=1, column=13).value = "AU12_r"
    sheet.cell(row=1, column=14).value = "AU15_l"
    sheet.cell(row=1, column=15).value = "AU15_r"
    sheet.cell(row=1, column=16).value = "AU25_t"
    sheet.cell(row=1, column=17).value = "AU26_t"
    sheet.cell(row=1, column=18).value = "Eyebrow distance"
    sheet.cell(row=1, column=19).value = "Drowsiness"
    sheet.cell(row=1, column=20).value = "EAR"
    sheet.cell(row=1, column=21).value = "Average Minimum Pitch"
    sheet.cell(row=1, column=22).value = "Average Minimum Yaw"

    df = pd.read_csv('benchmark_geometric.csv')
    sheet.cell(row=2, column=1).value = "Threshold"
    sheet.cell(row=2, column=2).value = df.iloc[0][1]
    sheet.cell(row=2, column=3).value = df.iloc[0][1]
    sheet.cell(row=2, column=4).value = df.iloc[1][1]
    sheet.cell(row=2, column=5).value = df.iloc[1][1]
    sheet.cell(row=2, column=6).value = df.iloc[2][1]
    sheet.cell(row=2, column=7).value = df.iloc[3][1]
    sheet.cell(row=2, column=8).value = df.iloc[4][1]
    sheet.cell(row=2, column=9).value = df.iloc[5][1]
    sheet.cell(row=2, column=10).value = df.iloc[5][1]
    sheet.cell(row=2, column=11).value = df.iloc[6][1]
    sheet.cell(row=2, column=12).value = df.iloc[7][1]
    sheet.cell(row=2, column=13).value = df.iloc[7][1]
    sheet.cell(row=2, column=14).value = df.iloc[8][1]
    sheet.cell(row=2, column=15).value = df.iloc[8][1]
    sheet.cell(row=2, column=16).value = df.iloc[9][1]
    sheet.cell(row=2, column=17).value = df.iloc[10][1]
    sheet.cell(row=2, column=18).value = df.iloc[13][1]
    sheet.cell(row=2, column=19).value = df.iloc[14][1]
    sheet.cell(row=2, column=20).value = "na"
    sheet.cell(row=2, column=21).value = "na"
    sheet.cell(row=2, column=22).value = "na"

    df = pd.read_excel(filename, "raw_data")
    neutral_frames = []
    neutral = 0
    starting_frames = []
    for i in tqdm(range(df.shape[0])):
        sheet.cell(row=4+i, column=1).value = i+1
        if df.loc[i, "Landmark_22_x"] == 0:
            continue

        a = (df.loc[i, "Landmark_22_x"], df.loc[i, "Landmark_22_y"])
        b = (df.loc[i, "Landmark_41_x"], df.loc[i, "Landmark_41_y"])
        AU1_l = np.linalg.norm(np.array(a)-np.array(b))
        a = (df.loc[i, "Landmark_23_x"], df.loc[i, "Landmark_23_y"])
        b = (df.loc[i, "Landmark_48_x"], df.loc[i, "Landmark_48_y"])
        AU1_r = np.linalg.norm(np.array(a)-np.array(b))
        sheet.cell(row=4+i, column=2).value = AU1_l
        sheet.cell(row=4+i, column=3).value = AU1_r

        a = (df.loc[i, "Landmark_20_x"], df.loc[i, "Landmark_20_y"])
        b = (df.loc[i, "Landmark_42_x"], df.loc[i, "Landmark_42_y"])
        AU2_l = np.linalg.norm(np.array(a)-np.array(b))
        a = (df.loc[i, "Landmark_25_x"], df.loc[i, "Landmark_25_y"])
        b = (df.loc[i, "Landmark_47_x"], df.loc[i, "Landmark_47_y"])
        AU2_r = np.linalg.norm(np.array(a)-np.array(b))
        sheet.cell(row=4+i, column=4).value = AU2_l
        sheet.cell(row=4+i, column=5).value = AU2_r

        a = (df.loc[i, "Landmark_22_x"], df.loc[i, "Landmark_22_y"])
        b = (df.loc[i, "Landmark_23_x"], df.loc[i, "Landmark_23_y"])
        AU4_t = np.linalg.norm(np.array(a)-np.array(b))
        sheet.cell(row=4+i, column=6).value = AU4_t

        AU5_t = df.loc[i, "EAR"]
        sheet.cell(row=4+i, column=7).value = AU5_t

        AU6_t = df.loc[i, "EAR"]
        sheet.cell(row=4+i, column=8).value = AU6_t

        a = (df.loc[i, "Landmark_22_x"], df.loc[i, "Landmark_22_y"])
        b = (df.loc[i, "Landmark_29_x"], df.loc[i, "Landmark_29_y"])
        AU9_l = np.linalg.norm(np.array(a)-np.array(b))
        a = (df.loc[i, "Landmark_23_x"], df.loc[i, "Landmark_23_y"])
        b = (df.loc[i, "Landmark_29_x"], df.loc[i, "Landmark_29_y"])
        AU9_r = np.linalg.norm(np.array(a)-np.array(b))
        sheet.cell(row=4+i, column=9).value = AU9_l
        sheet.cell(row=4+i, column=10).value = AU9_r

        a = (df.loc[i, "Landmark_49_x"], df.loc[i, "Landmark_49_y"])
        b = (df.loc[i, "Landmark_55_x"], df.loc[i, "Landmark_55_y"])
        AU12_t = np.linalg.norm(np.array(a)-np.array(b))
        AU12_l = df.loc[i, "Landmark_49_y"] - df.loc[i, "Landmark_34_y"]
        AU12_r = df.loc[i, "Landmark_55_y"] - df.loc[i, "Landmark_34_y"]
        sheet.cell(row=4+i, column=11).value = AU12_t
        sheet.cell(row=4+i, column=12).value = AU12_l
        sheet.cell(row=4+i, column=13).value = AU12_r

        a = (df.loc[i, "Landmark_49_x"], df.loc[i, "Landmark_49_y"])
        b = (df.loc[i, "Landmark_8_x"], df.loc[i, "Landmark_8_y"])
        AU15_l = np.linalg.norm(np.array(a)-np.array(b))
        a = (df.loc[i, "Landmark_55_x"], df.loc[i, "Landmark_55_y"])
        b = (df.loc[i, "Landmark_10_x"], df.loc[i, "Landmark_10_y"])
        AU15_r = np.linalg.norm(np.array(a)-np.array(b))
        sheet.cell(row=4+i, column=14).value = AU15_l
        sheet.cell(row=4+i, column=15).value = AU15_r

        a = (df.loc[i, "Landmark_63_x"], df.loc[i, "Landmark_63_y"])
        b = (df.loc[i, "Landmark_67_x"], df.loc[i, "Landmark_67_y"])
        AU25_t = np.linalg.norm(np.array(a)-np.array(b))
        sheet.cell(row=4+i, column=16).value = AU25_t

        a = (df.loc[i, "Landmark_34_x"], df.loc[i, "Landmark_34_y"])
        b = (df.loc[i, "Landmark_9_x"], df.loc[i, "Landmark_9_y"])
        AU26_t = np.linalg.norm(np.array(a)-np.array(b))
        sheet.cell(row=4+i, column=17).value = AU26_t

        a = (df.loc[i, "Landmark_20_x"], df.loc[i, "Landmark_20_y"])
        b = (df.loc[i, "Landmark_38_x"], df.loc[i, "Landmark_38_y"])
        eyebrow_dist_l = np.linalg.norm(np.array(a)-np.array(b))
        a = (df.loc[i, "Landmark_25_x"], df.loc[i, "Landmark_25_y"])
        b = (df.loc[i, "Landmark_45_x"], df.loc[i, "Landmark_45_y"])
        eyebrow_dist_r = np.linalg.norm(np.array(a)-np.array(b))
        sheet.cell(row=4+i, column=18).value = min(eyebrow_dist_l, eyebrow_dist_r)

        sheet.cell(row=4+i, column=19).value = df.loc[i, "EAR"]
        sheet.cell(row=4+i, column=20).value = df.loc[i, "EAR"]
        sheet.cell(row=4+i, column=21).value = df.loc[i, "Pitch_Min"]
        sheet.cell(row=4+i, column=22).value = df.loc[i, "Yaw_Min"]

        if neutral < 50:
            if (df.loc[i, "AU1":"AU26"] <= 0.2).all():
                landmarks = []
                landmarks.extend([AU1_l, AU1_r, AU2_l, AU2_r, AU4_t, AU5_t,
                                  AU6_t, AU9_l, AU9_r, AU12_t, AU12_l, AU12_r,
                                  AU15_l, AU15_r, AU25_t, AU26_t,
                                  min(eyebrow_dist_l, eyebrow_dist_r),
                                  df.loc[i, "EAR"], df.loc[i, "EAR"]])
                neutral_frames.append(landmarks)
                neutral += 1
        if 50 < i <= 150:
            landmarks = []
            landmarks.extend([AU1_l, AU1_r, AU2_l, AU2_r, AU4_t, AU5_t,
                              AU6_t, AU9_l, AU9_r, AU12_t, AU12_l, AU12_r,
                              AU15_l, AU15_r, AU25_t, AU26_t,
                              min(eyebrow_dist_l, eyebrow_dist_r),
                              df.loc[i, "EAR"], df.loc[i, "EAR"]])
            starting_frames.append(landmarks)

    if neutral >= 50:
        averages = np.mean(neutral_frames, 0)
    else:
        averages = np.mean(starting_frames, 0)
    sheet.cell(row=3, column=1).value = "Neutral"
    sheet.cell(row=3, column=2).value = averages[0]
    sheet.cell(row=3, column=3).value = averages[1]
    sheet.cell(row=3, column=4).value = averages[2]
    sheet.cell(row=3, column=5).value = averages[3]
    sheet.cell(row=3, column=6).value = averages[4]
    sheet.cell(row=3, column=7).value = averages[5]
    sheet.cell(row=3, column=8).value = averages[6]
    sheet.cell(row=3, column=9).value = averages[7]
    sheet.cell(row=3, column=10).value = averages[8]
    sheet.cell(row=3, column=11).value = averages[9]
    sheet.cell(row=3, column=12).value = averages[10]
    sheet.cell(row=3, column=13).value = averages[11]
    sheet.cell(row=3, column=14).value = averages[12]
    sheet.cell(row=3, column=15).value = averages[13]
    sheet.cell(row=3, column=16).value = averages[14]
    sheet.cell(row=3, column=17).value = averages[15]
    sheet.cell(row=3, column=18).value = averages[16]
    sheet.cell(row=3, column=19).value = averages[17]
    sheet.cell(row=3, column=20).value = averages[18]
    sheet.cell(row=3, column=21).value = np.mean(df.loc[:, "Pitch_Min"])
    sheet.cell(row=3, column=22).value = np.mean(df.loc[:, "Yaw_Min"])

    book.save(filename)

    #Tweak raw data using benchmarks
    df = pd.read_excel(filename, "geometric_calcs")
    min_confidence = 0.5
    book = openpyxl.load_workbook(filename)
    raw_data_sheet = book["raw_data"]
    if "raw_data Copy" in book.sheetnames:
            raw_data_copy_sheet = book["raw_data Copy"]
            book.remove(raw_data_copy_sheet)
    book.copy_worksheet(raw_data_sheet)
    raw_data_copy_sheet = book["raw_data Copy"]
    print("Tweaking raw data...")

    df2 = pd.read_excel(filename, "raw_data")
    #check AU1 & AU4 for % over 50%, then change reduction value
    threshold_for_AU_activation_in_total_video = 0.3
    reduction = 5
    AU1_reduction = 3
    AU4_reduction = 3
    amount_of_AU1_activated = (df2[df2.loc[:, "AU1"] >= min_confidence].shape[0]) / df2.shape[0]
    if amount_of_AU1_activated >= threshold_for_AU_activation_in_total_video:
        AU1_reduction = 5
    amount_of_AU4_activated = (df2[df2.loc[:, "AU4"] >= min_confidence].shape[0]) / df2.shape[0]
    if amount_of_AU4_activated >= threshold_for_AU_activation_in_total_video:
        AU4_reduction = 5
    for i in tqdm(range(1, len(raw_data_copy_sheet['A']))):
        if raw_data_copy_sheet.cell(row=i+1, column=20).value == 0:
            continue

        avg_pitch_min = df.loc[1, "Average Minimum Pitch"]
        pitch_min = raw_data_copy_sheet.cell(row=i+1, column=13).value
        pitch_delta = abs(avg_pitch_min - pitch_min)
        avg_yaw_min = df.loc[1, "Average Minimum Yaw"]
        yaw_min = raw_data_copy_sheet.cell(row=i+1, column=15).value
        yaw_delta = abs(avg_yaw_min - yaw_min)
        lower_threshold_yaw = (df.loc[1, "Average Minimum Yaw"]) - 10
        upper_threshold_yaw = (df.loc[1, "Average Minimum Yaw"]) + 10
        #Check if head position is too different from normal position and make all AUs 0s
        if pitch_delta >= 20 or yaw_delta >= 20:
            for j in range(2, 13):
                raw_data_copy_sheet.cell(row=i+1, column=j).value = 0
            continue

        AU1 = raw_data_copy_sheet.cell(row=i+1, column=2).value
        AU2a = raw_data_copy_sheet.cell(row=i+1, column=3).value
        AU2b = raw_data_copy_sheet.cell(row=i+1, column=4).value
        AU4 = raw_data_copy_sheet.cell(row=i+1, column=5).value
        AU5 = raw_data_copy_sheet.cell(row=i+1, column=6).value
        AU6 = raw_data_copy_sheet.cell(row=i+1, column=7).value
        AU9 = raw_data_copy_sheet.cell(row=i+1, column=8).value
        AU12 = raw_data_copy_sheet.cell(row=i+1, column=9).value
        AU15 = raw_data_copy_sheet.cell(row=i+1, column=10).value
        AU25 = raw_data_copy_sheet.cell(row=i+1, column=11).value
        AU26 = raw_data_copy_sheet.cell(row=i+1, column=12).value

        if AU1 >= min_confidence:
            if ((df.loc[i+1, "AU1_l"] - df.loc[1, "AU1_l"]) < df.loc[0, "AU1_l"]) and ((df.loc[i+1, "AU1_r"] - df.loc[1, "AU1_r"]) < df.loc[0, "AU1_r"]):
                if ((df.loc[1, "AU4_t"] - df.loc[i+1, "AU4_t"]) < df.loc[0, "AU4_t"]):
                    raw_data_copy_sheet.cell(row=i+1, column=2).value = AU1/reduction

        if AU2a >= min_confidence:
            if ((df.loc[i+1, "AU2_l"] - df.loc[1, "AU2_l"]) < df.loc[0, "AU2_l"]) and ((df.loc[i+1, "AU2_r"] - df.loc[1, "AU2_r"]) < df.loc[0, "AU2_r"]):
                if ((df.loc[1, "AU4_t"] - df.loc[i+1, "AU4_t"]) < df.loc[0, "AU4_t"]):
                    raw_data_copy_sheet.cell(row=i+1, column=3).value = AU2a/reduction

        if AU2b >= min_confidence:
            if ((df.loc[i+1, "AU2_l"] - df.loc[1, "AU2_l"]) < df.loc[0, "AU2_l"]) and ((df.loc[i+1, "AU2_r"] - df.loc[1, "AU2_r"]) < df.loc[0, "AU2_r"]):
                if ((df.loc[1, "AU4_t"] - df.loc[i+1, "AU4_t"]) < df.loc[0, "AU4_t"]):
                    raw_data_copy_sheet.cell(row=i+1, column=4).value = AU2b/reduction

        if AU4 >= min_confidence:
            if ((df.loc[1, "AU4_t"] - df.loc[i+1, "AU4_t"]) < df.loc[0, "AU4_t"]):
                raw_data_copy_sheet.cell(row=i+1, column=5).value = AU4/reduction

#        if AU5 >= min_confidence:
#            if ((df.loc[i+1, "AU5_t"] - df.loc[1, "AU5_t"]) < df.loc[0, "AU5_t"]):
#                raw_data_copy_sheet.cell(row=i+1, column=6).value = AU5/reduction

        if AU6 >= min_confidence:
            if ((df.loc[1, "AU6_t"] - df.loc[i+1, "AU6_t"]) < df.loc[0, "AU6_t"]):
                raw_data_copy_sheet.cell(row=i+1, column=7).value = AU6/reduction

        if AU9 >= min_confidence:
            if ((df.loc[1, "AU9_l"] - df.loc[i+1, "AU9_l"]) < df.loc[0, "AU9_l"]):
                raw_data_copy_sheet.cell(row=i+1, column=8).value = AU9/reduction
            elif ((df.loc[1, "AU9_r"] - df.loc[i+1, "AU9_r"]) < df.loc[0, "AU9_r"]):
                raw_data_copy_sheet.cell(row=i+1, column=8).value = AU9/reduction
            elif (yaw_min < lower_threshold_yaw) or (yaw_min > upper_threshold_yaw):
                raw_data_copy_sheet.cell(row=i+1, column=8).value = AU9/reduction

        if AU12 >= min_confidence:
            if (yaw_min < lower_threshold_yaw) or (yaw_min > upper_threshold_yaw):
                raw_data_copy_sheet.cell(row=i+1, column=9).value = AU12/reduction

        if AU15 >= min_confidence:
            if ((df.loc[1, "AU15_l"] - df.loc[i+1, "AU15_l"]) < df.loc[0, "AU15_l"]):
                raw_data_copy_sheet.cell(row=i+1, column=10).value = AU15/reduction
            elif ((df.loc[1, "AU15_r"] - df.loc[i+1, "AU15_r"]) < df.loc[0, "AU15_r"]):
                raw_data_copy_sheet.cell(row=i+1, column=10).value = AU15/reduction

        if AU25 >= min_confidence:
            if ((df.loc[i+1, "AU25_t"] - df.loc[1, "AU25_t"]) < df.loc[0, "AU25_t"]):
                raw_data_copy_sheet.cell(row=i+1, column=11).value = AU25/reduction

        if AU26 >= min_confidence:
            if ((df.loc[i+1, "AU26_t"] - df.loc[1, "AU26_t"]) < df.loc[0, "AU26_t"]):
                raw_data_copy_sheet.cell(row=i+1, column=12).value = AU26/reduction

    book.save(filename)
    df = pd.read_excel(filename, "raw_data Copy")
    df2 = pd.read_excel(filename, "geometric_calcs")

    #Process data for video
    print("Creating new processed data sheet...")
    book = openpyxl.load_workbook(filename)
    if "processed_data" in book.sheetnames:
            sheet = book["processed_data"]
            book.remove(sheet)
    sheet = book.create_sheet(title="processed_data")
    sheet.cell(row=1, column=1).value = "Time in Seconds"
    sheet.cell(row=1, column=2).value = "AU1"
    sheet.cell(row=1, column=3).value = "AU2"
    sheet.cell(row=1, column=4).value = "AU4"
    sheet.cell(row=1, column=5).value = "AU5"
    sheet.cell(row=1, column=6).value = "AU6"
    sheet.cell(row=1, column=7).value = "AU9"
    sheet.cell(row=1, column=8).value = "AU12"
    sheet.cell(row=1, column=9).value = "AU15"
    sheet.cell(row=1, column=10).value = "AU25"
    sheet.cell(row=1, column=11).value = "AU26"
    sheet.cell(row=1, column=12).value = "EAR"
    sheet.cell(row=1, column=13).value = "EAR Activated"
    sheet.cell(row=1, column=14).value = "Valence"
    sheet.cell(row=1, column=15).value = "Arousal"

    no_of_frames = len(df)
    no_of_seconds = no_of_frames//10
    leftover_frames = no_of_frames % 10
    extra_half_seconds = 0
    if 0 < leftover_frames <= 5:
        extra_half_seconds = 1
    elif leftover_frames > 5:
        extra_half_seconds = 2

    for i in range(1, 16):
        sheet.cell(row=2, column=i).value = 0.00
    sheet.cell(row=2, column=13).value = "No"

    #Take AU2 depending on distance between eyes and eyebrows
    if df2.loc[1, "Eyebrow distance"] > df2.loc[0, "Eyebrow distance"]:
        AU2_col = 2
    else:
        AU2_col = 3

    print("Processing data...")
    for i in tqdm(range(1, no_of_seconds*2 + extra_half_seconds + 1)):
        second = i / 2
        sheet.cell(row=i+2, column=1).value = second
        frame_min = (second - 0.5) * 10
        frame_max = second * 10 - 1
        for j in range(1, 12):
            if j == 1:
                sheet.cell(row=i+2, column=j+1).value = np.mean(df.loc[frame_min:frame_max, df.columns[j]])
            elif j == 2:
                sheet.cell(row=i+2, column=j+1).value = np.mean(df.loc[frame_min:frame_max, df.columns[AU2_col]])
            elif j == 11:
                sheet.cell(row=i+2, column=j+1).value = np.mean(df.loc[frame_min:frame_max, df.columns[154]])
            else:
                sheet.cell(row=i+2, column=j+1).value = np.mean(df.loc[frame_min:frame_max, df.columns[j+1]])
        #Check if the average EAR ratio over the last 5 seconds is less than threshold to see if they are sleepy
        #But not if cheek raise!
        if (df2.loc[1, "Drowsiness"] - np.mean(df.loc[max(0,frame_max-50):frame_max, df.columns[154]])) > df2.loc[0, "Drowsiness"] and i != 1 and np.mean(df.loc[frame_min:frame_max, df.columns[6]]) < min_confidence:
                sheet.cell(row=i+2, column=13).value = 'Yes'
        else:
            sheet.cell(row=i+2, column=13).value = 'No'

    print("Finished processing AU data...")
    book.save(filename)
    print("Predicting valence and arousal...")

    #Predicting valence and arousal
    book = openpyxl.load_workbook(filename)
    sheet = book["raw_data Copy"]
    df = pd.read_excel(filename, "raw_data Copy")
    # Importing data for valence and arousal prediction
    df = df.iloc[:,1:12]
    df = df.dropna()
    data = df.values
    x1 = data[:,0]
    x2 = data[:,AU2_col-1]
    x3 = data[:,3:11]
    x = np.concatenate((x1.reshape(-1,1), x2.reshape(-1,1), x3), axis=1)
    x = x.astype('float32')
    #Valence predictions
    model_valence = joblib.load('models/Valence/valence.pkl')
    val_preds = model_valence.predict(x)
    val_preds = np.clip(val_preds, a_min = -1, a_max = 1)
    val_mean = np.mean(val_preds)
    #Valence calibration
    val_cal = val_preds - val_mean
    val_cal = np.clip(val_cal, a_min = -1, a_max = 1)
    #Valence scaled
    val_scaled = (val_cal + 1) * 10 / 2
    sheet.cell(row=1, column=156).value = "Valence"
    sheet.cell(row=1, column=157).value = "Valence calibrated"
    sheet.cell(row=1, column=158).value = "Valence scaled"
    for i in range(len(val_preds)):
        sheet.cell(row=i+2, column=156).value = val_preds[i]
        sheet.cell(row=i+2, column=157).value = val_cal[i]
        sheet.cell(row=i+2, column=158).value = val_scaled[i]
    #Arousal predictions
    model_arousal = joblib.load('models/Arousal/arousal.pkl')
    ars_preds = model_arousal.predict(x)
    ars_preds = np.clip(ars_preds, a_min = -1, a_max = 1)
    ars_mean = np.mean(ars_preds)
    #Arousal calibration
    ars_cal = ars_preds - ars_mean
    ars_cal = np.clip(ars_cal, a_min = -1, a_max = 1)
    #Arousal scaled
    ars_scaled = (ars_cal + 1) * 10 / 2
    sheet.cell(row=1, column=158).value = "Arousal"
    sheet.cell(row=1, column=159).value = "Arousal calibrated"
    sheet.cell(row=1, column=160).value = "Arousal scaled"
    for i in range(len(ars_preds)):
        sheet.cell(row=i+2, column=158).value = ars_preds[i]
        sheet.cell(row=i+2, column=159).value = ars_cal[i]
        sheet.cell(row=i+2, column=160).value = ars_scaled[i]
    #Engagement Level
    sheet.cell(row=1, column=161).value = "Engagement"
    for i in range(len(ars_preds)):
        sheet.cell(row=i+2, column=161).value = (ars_scaled[i] * (2/3)) + (val_scaled[i] * (1/3))
    print("Saving file...")
    book.save(filename)

def create_charts_for_report(filename):
    df = pd.read_excel(filename, "raw_data Copy")
    filename, ext = os.path.splitext(filename)
    #Setting standard parameters such as pixel size and dpi
    my_dpi = 120
    plt.rcParams.update({'font.size': 7})
    width = 945
    height = 355
    #Create Valence chart
    val_answer_df = df[['Frame','Valence']]
    ars_answers = [0, 0, 0, 0, 0] #Example
    val_answer_df.loc[0:100, 'Valence'] = 0
    for i in range(5):
        if i == 0:
            val_answer_df.loc[100:830,'Valence'] = val_answers[i]
        elif i == 1:
            val_answer_df.loc[830:1450,'Valence'] = val_answers[i]
        elif i == 2:
            val_answer_df.loc[1450:1870,'Valence'] = val_answers[i]
        elif i == 3:
            val_answer_df.loc[1870:2240,'Valence'] = val_answers[i]
        else:
            val_answer_df.loc[2240:,'Valence'] = val_answers[i]

    valence = 'Valence calibrated'
    val_answers_plot = False
    video_line_color = 'salmon'

    plt.figure(figsize=(width/my_dpi, height/my_dpi), dpi=my_dpi)
    val_df = pd.DataFrame(data=df[valence],columns=[valence])
    val_exp = val_df[valence].ewm(span=50, adjust=False).mean()
    plt.plot(list(val_df.index/600), val_exp)
    if val_answers_plot == True:
        plt.plot(list(val_df.index/600), val_answer_df['Valence'], color='g')
        video_line_color = 'gray'
    axes = plt.gca()
    axes.set_ylim([-1,1])
    axes.set_xlim([0, (df['Frame']/600).max()])
    plt.xticks(np.arange(0, 5, 0.5))
    plt.grid(True)
    plt.axvline(x=(10/60), color=video_line_color, linewidth=1, linestyle='dashed')
    plt.axvline(x=(83/60), color=video_line_color, linewidth=1, linestyle='dashed')
    plt.axvline(x=(145/60), color=video_line_color, linewidth=1, linestyle='dashed')
    plt.axvline(x=(187/60), color=video_line_color, linewidth=1, linestyle='dashed')
    plt.axvline(x=(224/60), color=video_line_color, linewidth=1, linestyle='dashed')
    plt.text((8/60),-1.1,'V1',rotation=0, weight='bold')
    plt.text((78/60),-1.1,'V2',rotation=0, weight='bold')
    plt.text((139/60),-1.1,'V3',rotation=0, weight='bold')
    plt.text((185/60),-1.1,'V4',rotation=0, weight='bold')
    plt.text((221/60),-1.1,'V5',rotation=0, weight='bold')
    plt.axhline(y=0, color='gray', linewidth=1, linestyle='dashed')
    plt.ylabel('Valence Level')
    plt.xlabel('Time in minutes')
    plt.show()

    #Create Arousal chart
    ars_answer_df = df[['Frame','Arousal']]
    ars_answers = [0, 0, 0, 0, 0] #Example
    ars_answer_df.loc[0:100, 'Arousal'] = 0
    for i in range(5):
        if i == 0:
            ars_answer_df.loc[100:830,'Arousal'] = ars_answers[i]
        elif i == 1:
            ars_answer_df.loc[830:1450,'Arousal'] = ars_answers[i]
        elif i == 2:
            ars_answer_df.loc[1450:1870,'Arousal'] = ars_answers[i]
        elif i == 3:
            ars_answer_df.loc[1870:2240,'Arousal'] = ars_answers[i]
        else:
            ars_answer_df.loc[2240:,'Arousal'] = ars_answers[i]

    arousal = 'Arousal calibrated'
    ars_answers_plot = False
    video_line_color = 'salmon'

    plt.figure(figsize=(width/my_dpi, height/my_dpi), dpi=my_dpi)
    ars_df = pd.DataFrame(data=df[arousal],columns=[arousal])
    ars_exp = ars_df[arousal].ewm(span=50, adjust=False).mean()
    plt.plot(list(ars_df.index/600), ars_exp)
    if ars_answers_plot == True:
        plt.plot(list(ars_df.index/600), ars_answer_df['Arousal'], color='g')
        video_line_color = 'gray'
    axes = plt.gca()
    axes.set_ylim([-1,1])
    axes.set_xlim([0, (df['Frame']/600).max()])
    plt.xticks(np.arange(0, 5, 0.5))
    plt.grid(True)
    plt.axvline(x=(10/60), color=video_line_color, linewidth=1, linestyle='dashed')
    plt.axvline(x=(83/60), color=video_line_color, linewidth=1, linestyle='dashed')
    plt.axvline(x=(145/60), color=video_line_color, linewidth=1, linestyle='dashed')
    plt.axvline(x=(187/60), color=video_line_color, linewidth=1, linestyle='dashed')
    plt.axvline(x=(224/60), color=video_line_color, linewidth=1, linestyle='dashed')
    plt.text((8/60),-1.1,'V1',rotation=0, weight='bold')
    plt.text((78/60),-1.1,'V2',rotation=0, weight='bold')
    plt.text((139/60),-1.1,'V3',rotation=0, weight='bold')
    plt.text((185/60),-1.1,'V4',rotation=0, weight='bold')
    plt.text((221/60),-1.1,'V5',rotation=0, weight='bold')
    plt.axhline(y=0, color='gray', linewidth=1, linestyle='dashed')
    plt.ylabel('Arousal Level')
    plt.xlabel('Time in minutes')
    plt.show()

    #Create Engagement chart from actual valence and arousals
    val_answers_scaled = (val_answer_df['Valence'] + 1) * 10 / 2
    ars_answers_scaled = (ars_answer_df['Arousal'] + 1) * 10 / 2
    eng_answers = val_answers_scaled.iloc[:] * (1/3) + ars_answers_scaled.iloc[:] * (2/3)

    eng_answers_plot = False
    video_line_color = 'salmon'

    plt.figure(figsize=(width/my_dpi, height/my_dpi), dpi=my_dpi)
    engagement = df.loc[:,'Engagement']
    engagement = pd.DataFrame(data=engagement,columns=['Engagement'])
    eng_exp = engagement['Engagement'].ewm(span=50, adjust=False).mean()
    plt.plot(list(engagement.index/600), eng_exp)
    y_lim = [3,7]
    y_ticks = [3,4,5,6,7]
    video_label_x_axis = 2.80
    if eng_answers_plot == True:
        plt.plot(list(engagement.index/600), eng_answers.iloc[:], color='g')
        video_line_color = 'gray'
        y_lim = [0,10]
        y_ticks = [0,1,2,3,4,5,6,7,8,9,10]
        video_label_x_axis = -0.5
    axes.set_ylim(y_lim)
    plt.yticks(y_ticks)
    axes = plt.gca()
    axes.set_xlim([0, (df['Frame']/600).max()])
    plt.grid(True)
    plt.axvline(x=(10/60), color=video_line_color, linewidth=1, linestyle='dashed')
    plt.axvline(x=(83/60), color=video_line_color, linewidth=1, linestyle='dashed')
    plt.axvline(x=(145/60), color=video_line_color, linewidth=1, linestyle='dashed')
    plt.axvline(x=(187/60), color=video_line_color, linewidth=1, linestyle='dashed')
    plt.axvline(x=(224/60), color=video_line_color, linewidth=1, linestyle='dashed')
    plt.text((8/60),video_label_x_axis,'V1',rotation=0, weight='bold')
    plt.text((78/60),video_label_x_axis,'V2',rotation=0, weight='bold')
    plt.text((139/60),video_label_x_axis,'V3',rotation=0, weight='bold')
    plt.text((185/60),video_label_x_axis,'V4',rotation=0, weight='bold')
    plt.text((221/60),video_label_x_axis,'V5',rotation=0, weight='bold')
    plt.axhline(y=5, color='gray', linewidth=1, linestyle='dashed')
    plt.xticks(np.arange(0, 5, 0.5))
    plt.gca().yaxis.set_major_formatter(StrMethodFormatter('{x:,.2f}'))
    plt.ylabel('Learner Engagement Level')
    plt.show()

def create_valence_and_arousal_charts(filename):
    df = pd.read_excel(filename, "raw_data Copy")
    filename, ext = os.path.splitext(filename)
    #Setting standard parameters such as pixel size and dpi
    my_dpi = 120
    plt.rcParams.update({'font.size': 7})
    width = 945
    height = 144
    #Create Valence chart
    plt.figure(figsize=(width/my_dpi, height/my_dpi), dpi=my_dpi)
    val_df = pd.DataFrame(data=df['Valence calibrated'],columns=['Valence calibrated'])
    val_exp = val_df['Valence calibrated'].ewm(span=50, adjust=False).mean()
    plt.plot(list(val_df.index/600), val_exp)
    axes = plt.gca()
    axes.set_ylim([-1,1])
    axes.set_xlim([0, (df['Frame']/600).max()])
    axes.get_xaxis().set_visible(False)
    plt.grid(True)
    plt.ylabel('Valence Level')
    plt.savefig(filename + '_valence.png', bbox_inches='tight',pad_inches = 0, dpi=my_dpi)

    #Create Arousal chart
    plt.figure(figsize=(width/my_dpi, height/my_dpi), dpi=my_dpi)
    ars_df = pd.DataFrame(data=df['Arousal calibrated'],columns=['Arousal calibrated'])
    ars_exp = ars_df['Arousal calibrated'].ewm(span=50, adjust=False).mean()
    plt.plot(list(ars_df.index/600), ars_exp)
    axes = plt.gca()
    axes.set_ylim([-1,1])
    axes.set_xlim([0, (df['Frame']/600).max()])
    axes.get_xaxis().set_visible(False)
    plt.grid(True)
    plt.ylabel('Arousal Level')
    plt.savefig(filename + '_arousal.png', bbox_inches='tight',pad_inches = 0, dpi=my_dpi)

    #Setting standard parameters such as pixel size and dpi
    my_dpi = 120
    plt.rcParams.update({'font.size': 7})
    width = 940
    height = 355
    #Create Engagement chart from actual valence and arousals
    plt.figure(figsize=(width/my_dpi, height/my_dpi), dpi=my_dpi)
    engagement = df.loc[:,'Engagement']
    engagement = pd.DataFrame(data=engagement,columns=['Engagement'])
    eng_exp = engagement['Engagement'].ewm(span=50, adjust=False).mean()
    plt.plot(list(engagement.index/600), eng_exp)
    axes = plt.gca()
#    axes.set_ylim([0,10]) #To expand chart to 0 to 10 on yaxis
#    plt.yticks(np.arange(0, 11, 2.0)) #To expand chart to 0 to 10 on yaxis
    axes.set_ylim([3,7]) #To shrink chart to 3 to 7 on yaxis
    plt.yticks([3,4,5,6,7]) #To shrink chart to 3 to 7 on yaxis
    axes.set_xlim([0, (df['Frame']/600).max()])
    axes.get_xaxis().set_visible(False)
    plt.grid(True)
    plt.axhline(y=5, color='gray', linewidth=1, linestyle='dashed')
    plt.ylabel('Learner Engagement Level')
    plt.savefig(filename + '_engagement.png', bbox_inches='tight',pad_inches = 0, dpi=my_dpi)

if __name__ == "__main__":
    filename =  sys.argv[1] #sys.argv[1]
    print("Loading models...")
    predictor = Video_Predictor()
    print("Processing video...")
    predictions = predictor.process_video(filename)
    save_to_excel(predictions, filename)